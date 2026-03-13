import argparse
import copy
import csv
import gc
import hashlib
import json
import multiprocessing as mp
import os
from pathlib import Path
import pickle
import re
import time
import traceback
from typing import Any, Dict, List, Optional, Tuple

import torch
import torch.nn as nn
from torch.utils.data import DataLoader, Dataset
import numpy as np
from dotmap import DotMap
import wandb
from wandb.wandb_run import Run
import openpyxl
from openpyxl.styles import Alignment
from PIL import Image
from collections.abc import Mapping
from contextlib import nullcontext
from datetime import datetime
from einops import rearrange
from sklearn.linear_model import Ridge, RidgeCV
from scipy import stats
from scipy.optimize import curve_fit
import inspect

from utils.torch_amp_compat import patch_cuda_amp_custom_autocast

patch_cuda_amp_custom_autocast(device_type="cuda")

from data import (
    LIVEDataset,
    CSIQDataset,
    TID2013Dataset,
    KADID10KDataset,
    FLIVEDataset,
    SPAQDataset,
)
from utils.utils import (
    PROJECT_ROOT,
    parse_command_line_args,
    merge_configs,
    parse_config,
    replace_none_string,
    prepare_wandb_config,
)
from models.associations import (
    covariance_energy_metric,
    covariance_rank_metric,
    feature_norm_metric,
    feature_spread_metric,
    mse_agreement_metric,
    normalized_spread_metric,
)
from models.simclr import SimCLR
from models.vicreg import Vicreg
import warnings
from joblib import Parallel, delayed

os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
try:
    mp.set_start_method("spawn", force=True)
except RuntimeError:
    pass

synthetic_datasets = ["live", "csiq", "tid2013", "kadid10k"]
authentic_datasets = ["flive", "spaq"]

try:
    torch.multiprocessing.set_sharing_strategy("file_system")
except RuntimeError:
    pass
try:
    mp.set_start_method("spawn", force=True)
except RuntimeError:
    pass


def _env_int(name: str, default: int) -> int:
    value = os.environ.get(name, None)
    if value is None:
        return int(default)
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def _env_bool(name: str, default: bool) -> bool:
    value = os.environ.get(name, None)
    if value is None:
        return bool(default)
    value = str(value).strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    return bool(default)


def _coerce_int(value, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def prepare_dataset(
    dataset_key: str,
    data_base_path: Path,
    crop_size: int,
    default_num_splits: int,
    fr_iqa: bool = False,
):
    dataset_key = dataset_key.lower()
    if dataset_key == "live":
        dataset = LIVEDataset(
            data_base_path / "LIVE",
            phase="all",
            crop_size=crop_size,
            fr_mode=fr_iqa,
        )
        dataset_num_splits = default_num_splits
        dataset_name = "LIVE"
    elif dataset_key == "csiq":
        dataset = CSIQDataset(
            data_base_path / "CSIQ",
            phase="all",
            crop_size=crop_size,
            fr_mode=fr_iqa,
        )
        dataset_num_splits = default_num_splits
        dataset_name = "CSIQ"
    elif dataset_key == "tid2013":
        dataset = TID2013Dataset(
            data_base_path / "TID2013",
            phase="all",
            crop_size=crop_size,
            fr_mode=fr_iqa,
        )
        dataset_num_splits = default_num_splits
        dataset_name = "TID2013"
    elif dataset_key == "kadid10k":
        dataset = KADID10KDataset(
            data_base_path / "KADID10K",
            phase="all",
            crop_size=crop_size,
            fr_mode=fr_iqa,
        )
        dataset_num_splits = default_num_splits
        dataset_name = "KADID-10K"
    elif dataset_key == "flive":
        if fr_iqa:
            raise ValueError(
                "FR-IQA is only supported for LIVE, CSIQ, TID2013, and KADID-10K."
            )
        dataset = FLIVEDataset(
            data_base_path / "FLIVE", phase="all", crop_size=crop_size
        )
        dataset_num_splits = 1
        dataset_name = "FLIVE"
    elif dataset_key == "spaq":
        if fr_iqa:
            raise ValueError(
                "FR-IQA is only supported for LIVE, CSIQ, TID2013, and KADID-10K."
            )
        dataset = SPAQDataset(
            data_base_path / "SPAQ", phase="all", crop_size=crop_size
        )
        dataset_num_splits = default_num_splits
        dataset_name = "SPAQ"
    else:
        raise ValueError(f"Dataset {dataset_key} not supported")
    return dataset, dataset_num_splits, dataset_name


def sanitize_identifier(value: str) -> str:
    value = value.replace("->", "_to_").replace(" ", "_")
    safe_chars = []
    for ch in value:
        if ch.isalnum() or ch in ("_", "-", "."):
            safe_chars.append(ch)
        else:
            safe_chars.append("_")
    sanitized = "".join(safe_chars)
    return sanitized.strip("_") or "result"


def sanitize_sheet_name(name: str) -> str:
    invalid_chars = set(r"[]:*?/\\")
    sanitized = "".join("_" if ch in invalid_chars else ch for ch in name)
    sanitized = sanitized.replace(" ", "_")
    sanitized = sanitized[:31] if len(sanitized) > 31 else sanitized
    sanitized = sanitized.strip()
    return sanitized or "Sheet"


def build_result_metadata(
    train_key: str,
    test_key: str,
    train_display: str,
    test_display: str,
):
    same_dataset = train_key == test_key
    display_name = (
        train_display if same_dataset else f"{train_display} -> {test_display}"
    )
    sheet_base = (
        train_display if same_dataset else f"{train_display}_to_{test_display}"
    )
    identifier_base = train_key if same_dataset else f"{train_key}_to_{test_key}"
    metadata = {
        "display_name": display_name,
        "train_dataset": train_key,
        "test_dataset": test_key,
        "sheet_name": sanitize_sheet_name(sheet_base),
        "file_stem": sanitize_identifier(identifier_base),
        "logger_suffix": sanitize_identifier(identifier_base),
    }
    return metadata


def safe_mean(values: List[float]) -> float:
    if not values:
        return float("nan")
    return float(np.mean(values))

def safe_max(values: List[float]) -> float:
    if not values:
        return float("nan")
    return float(np.max(values))


def format_float(value: float) -> str:
    if value is None:
        return "nan"
    if isinstance(value, (float, np.floating)):
        if np.isnan(value) or np.isinf(value):
            return "nan"
    return f"{value:.4f}"


def format_alpha(alpha_value) -> str:
    if isinstance(alpha_value, (float, np.floating)):
        return f"{alpha_value:.12g}"
    return str(alpha_value)


def resolve_eval_pin_memory(args=None, default: bool = False) -> bool:
    if args is None:
        return default
    test_cfg = getattr(args, "test", None)
    if test_cfg is None:
        return default
    value = (
        test_cfg.get("pin_memory", None)
        if hasattr(test_cfg, "get")
        else getattr(test_cfg, "pin_memory", None)
    )
    if value is None:
        return default
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "1", "yes", "on"}:
            return True
        if lowered in {"false", "0", "no", "off"}:
            return False
    return bool(value)


def _split_representation_views(features: torch.Tensor) -> Tuple[torch.Tensor, Optional[torch.Tensor]]:
    if features.ndim < 2:
        return features, None
    feat_dim = features.shape[1]
    if feat_dim == 0 or feat_dim % 2 != 0:
        return features, None
    half = feat_dim // 2
    return features[:, :half], features[:, half:]


def compute_vicreg_metrics(
    reps: np.ndarray, embs: np.ndarray, device: torch.device
) -> dict:
    metrics = {
        "corr": covariance_energy_metric,
        "std": feature_spread_metric,
        "l2norm": feature_norm_metric,
        "cov_rank": covariance_rank_metric,
        "nstd": normalized_spread_metric,
    }

    target_device = (
        "cuda" if device.type == "cuda" and torch.cuda.is_available() else "cpu"
    )
    inps = {
        "rep": torch.tensor(reps).to(target_device),
        "emb": torch.tensor(embs).to(target_device),
    }
    vicreg_metrics = {}
    for inp_n, inp_v in inps.items():
        for metric_n, metric_f in metrics.items():
            vicreg_metrics[f"{metric_n}_{inp_n}"] = metric_f(inp_v, inp_v)

        rank_key = f"cov_rank_{inp_n}"
        feat_dim = inp_v.shape[1] if inp_v.ndim >= 2 else 0
        if feat_dim:
            vicreg_metrics[f"nrank_{inp_n}"] = float(vicreg_metrics[rank_key]) / float(
                feat_dim
            )
        else:
            vicreg_metrics[f"nrank_{inp_n}"] = float("nan")

        view_a, view_b = _split_representation_views(inp_v)
        if view_b is None:
            vicreg_metrics[f"inv_{inp_n}"] = float("nan")
        else:
            vicreg_metrics[f"inv_{inp_n}"] = mse_agreement_metric(view_a, view_b)

    return vicreg_metrics


def _to_float_array(values):
    if isinstance(values, np.ndarray):
        return values.astype(np.float64, copy=False)
    return np.asarray(values, dtype=np.float64)


def convert_score_orientation(
    scores,
    from_type: str,
    to_type: str,
    score_range: Tuple[float, float],
):
    scores_arr = _to_float_array(scores)
    if from_type == to_type or scores_arr.size == 0:
        return scores_arr
    if score_range is None or len(score_range) != 2:
        return scores_arr
    min_val, max_val = score_range
    return (min_val + max_val) - scores_arr


def rescale_scores(
    scores,
    source_range: Tuple[float, float],
    target_range: Tuple[float, float],
):
    scores_arr = _to_float_array(scores)
    if scores_arr.size == 0:
        return scores_arr
    if (
        source_range is None
        or target_range is None
        or len(source_range) != 2
        or len(target_range) != 2
    ):
        return scores_arr
    src_min, src_max = source_range
    tgt_min, tgt_max = target_range
    if np.isclose(src_max, src_min):
        midpoint = tgt_min + (tgt_max - tgt_min) / 2.0
        return np.full_like(scores_arr, midpoint)
    normalized = (scores_arr - src_min) / (src_max - src_min)
    return normalized * (tgt_max - tgt_min) + tgt_min


def map_scores_between_datasets(
    scores,
    source_type: str,
    source_range: Tuple[float, float],
    target_type: str,
    target_range: Tuple[float, float],
):
    oriented = convert_score_orientation(scores, source_type, target_type, source_range)
    if source_range == target_range or target_range is None or source_range is None:
        return oriented
    return rescale_scores(oriented, source_range, target_range)


def resolve_fr_flag(args: DotMap) -> bool:
    return bool(args.get("fr_iqa", False) or args.test.get("fr_iqa", False))


def resolve_plcc_logistic(args: DotMap) -> bool:
    if args is None:
        return False
    if isinstance(args, Mapping) and "plcc_logistic" in args:
        return bool(args.get("plcc_logistic", False))
    test_section = getattr(args, "test", None)
    if test_section is None:
        return False
    return bool(test_section.get("plcc_logistic", False))


def resolve_fr_sanity_flag(args: DotMap) -> bool:
    return bool(
        args.get("fr_sanity_checks", False)
        or args.test.get("fr_sanity_checks", False)
    )


def resolve_fr_sanity_samples(args: DotMap, default: int = 64) -> int:
    value = args.get("fr_sanity_samples", None)
    if value is None:
        value = args.test.get("fr_sanity_samples", default)
    try:
        value = int(value)
    except (TypeError, ValueError):
        value = default
    return max(1, value)


def logistic_4p(x, beta1, beta2, beta3, beta4):
    beta4_safe = np.abs(beta4) + 1e-8
    z = -(x - beta3) / beta4_safe
    z = np.clip(z, -60.0, 60.0)
    return (beta1 - beta2) / (1.0 + np.exp(z)) + beta2


def apply_plcc_logistic_mapping(preds, targets):
    preds = np.asarray(preds, dtype=np.float64)
    targets = np.asarray(targets, dtype=np.float64)
    if preds.size < 2:
        return preds
    beta1 = float(np.max(targets))
    beta2 = float(np.min(targets))
    beta3 = float(np.mean(preds))
    beta4 = float(np.std(preds)) if float(np.std(preds)) > 0 else 1.0
    try:
        popt, _ = curve_fit(
            logistic_4p,
            preds,
            targets,
            p0=[beta1, beta2, beta3, beta4],
            maxfev=10000,
        )
        return logistic_4p(preds, *popt)
    except Exception:
        return preds


def describe_plcc_mapping(args: DotMap = None) -> Tuple[str, str]:
    if resolve_plcc_logistic(args):
        return "enabled", "logistic_4p+pearsonr"
    return "disabled", "pearsonr"


def check_reference_disjoint(dataset: Dataset, num_splits: int):
    if not hasattr(dataset, "ref_images"):
        return None
    if dataset.ref_images is None or len(dataset.ref_images) == 0:
        return None
    ref_ids = np.array([Path(p).stem for p in dataset.ref_images])
    for split_idx in range(num_splits):
        train_idx = dataset.get_split_indices(split_idx, "train")
        test_idx = dataset.get_split_indices(split_idx, "test")
        val_idx = dataset.get_split_indices(split_idx, "val")
        train_refs = set(ref_ids[train_idx])
        test_refs = set(ref_ids[test_idx])
        val_refs = set(ref_ids[val_idx])
        if train_refs & test_refs or train_refs & val_refs or test_refs & val_refs:
            return False
    return True


def describe_split_protocol(dataset: Dataset, num_splits: int, aggregation: str) -> str:
    total = len(dataset)
    train_idx = dataset.get_split_indices(0, "train")
    val_idx = dataset.get_split_indices(0, "val")
    test_idx = dataset.get_split_indices(0, "test")
    ratio = (
        f"{int(round(100 * len(train_idx) / max(total, 1)))}"
        f"/{int(round(100 * len(val_idx) / max(total, 1)))}"
        f"/{int(round(100 * len(test_idx) / max(total, 1)))}"
    )
    ref_disjoint = check_reference_disjoint(dataset, num_splits)
    ref_disjoint_str = (
        "unknown" if ref_disjoint is None else str(bool(ref_disjoint))
    )
    seeds = list(range(num_splits))
    return (
        f"ratio={ratio},seeds={seeds},aggregation={aggregation},ref_disjoint={ref_disjoint_str}"
    )


def resolve_checkpoint_label(args: DotMap, eval_type: str) -> str:
    if eval_type != "scratch":
        raise ValueError(f"Unsupported eval_type: {eval_type}")
    checkpoint_base_path = PROJECT_ROOT / "experiments"
    checkpoint_path = checkpoint_base_path / args.experiment_name / "pretrain"
    if checkpoint_path.exists():
        candidates = [ckpt for ckpt in checkpoint_path.glob("*.pth") if "best" in ckpt.name]
        if candidates:
            return candidates[0].name
    return args.experiment_name


_ALPHA_CACHE_FILE_VERSION = 1
_ALPHA_CACHE_MEMORY: Dict[str, Dict[str, Any]] = {}


def _normalize_none_like(value: Any) -> Any:
    if isinstance(value, str) and value.strip().lower() in {"", "none", "null"}:
        return None
    return value


def _resolve_eval_section(args: Optional[DotMap], phase: str) -> DotMap:
    if args is None:
        return DotMap(_dynamic=True)
    if phase == "cross" and hasattr(args, "test"):
        cross_cfg = getattr(args.test, "cross_eval", None)
        if isinstance(cross_cfg, Mapping):
            return DotMap(cross_cfg)
        if cross_cfg is not None:
            return cross_cfg
        return args.test
    if phase == "val" and hasattr(args, "validation"):
        return args.validation
    if hasattr(args, "test"):
        return args.test
    return DotMap(_dynamic=True)


def _resolve_bool_cfg(value: Any, default: bool) -> bool:
    value = _normalize_none_like(value)
    if value is None:
        return default
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"true", "false"}:
            return lowered == "true"
    return bool(value)


def _resolve_float_cfg(value: Any, default: float) -> float:
    value = _normalize_none_like(value)
    if value is None:
        return float(default)
    try:
        return float(value)
    except (TypeError, ValueError):
        return float(default)


def _resolve_int_cfg(value: Any, default: int) -> int:
    value = _normalize_none_like(value)
    if value is None:
        return int(default)
    try:
        return int(value)
    except (TypeError, ValueError):
        return int(default)


def _dump_final_eval_metrics(
    checkpoint_path: Path, logger: Optional[Run], payload: Dict[str, Any]
) -> None:
    normalized_payload: Dict[str, float] = {}
    for key, value in payload.items():
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            continue
        if np.isnan(numeric_value) or np.isinf(numeric_value):
            continue
        normalized_payload[key] = numeric_value

    if not normalized_payload:
        return

    output_paths = [checkpoint_path / "final_eval_metrics.json"]
    if logger is not None:
        try:
            output_paths.append(Path(logger.dir) / "final_eval_metrics.json")
        except Exception:
            pass

    for output_path in output_paths:
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(
                json.dumps(normalized_payload, indent=2, sort_keys=True),
                encoding="utf-8",
            )
        except Exception as exc:
            print(f"Warning: failed to write {output_path}: {exc}")


def _resolve_str_cfg(value: Any, default: Optional[str] = None) -> Optional[str]:
    value = _normalize_none_like(value)
    if value is None:
        return default
    return str(value).strip()


def _resolve_fast_mode_strategy(args: Optional[DotMap], phase: str) -> str:
    section = _resolve_eval_section(args, phase)
    test_section = getattr(args, "test", DotMap(_dynamic=True)) if args else DotMap(_dynamic=True)
    default_fast_mode = (
        _resolve_bool_cfg(test_section.get("fast_mode", False), False)
        if phase in {"test", "cross"}
        else False
    )
    fast_mode = _resolve_bool_cfg(section.get("fast_mode", default_fast_mode), default_fast_mode)
    strategy = _resolve_str_cfg(section.get("fast_mode_strategy", None), None)
    if strategy is not None:
        strategy = strategy.lower()
        if strategy == "fast":
            strategy = "legacy"
    if strategy is None:
        return "legacy" if fast_mode else "slow"
    if strategy not in {"slow", "legacy", "budget", "subsample", "cache"}:
        return "legacy" if fast_mode else "slow"
    return strategy


def _resolve_seed(args: Optional[DotMap], dataset_key: str, offset: int = 0) -> int:
    base_seed = 27
    if args is not None:
        try:
            base_seed = int(getattr(args, "seed", base_seed))
        except (TypeError, ValueError):
            base_seed = 27
    digest = hashlib.sha1(dataset_key.encode("utf-8")).hexdigest()[:8]
    dataset_seed = int(digest, 16)
    return int(base_seed + dataset_seed + offset)


def _resolve_budget_count(total_count: int, budget: Any, default_fraction: float) -> int:
    total_count = int(max(0, total_count))
    if total_count == 0:
        return 0
    budget = _normalize_none_like(budget)
    if budget is None:
        budget_value = float(default_fraction)
    else:
        try:
            budget_value = float(budget)
        except (TypeError, ValueError):
            budget_value = float(default_fraction)
    if budget_value <= 0:
        return total_count
    if budget_value <= 1:
        resolved = int(round(total_count * budget_value))
    else:
        resolved = int(round(budget_value))
    resolved = max(1, min(total_count, resolved))
    return resolved


def _sample_split_indices(num_splits: int, target_count: int, seed: int) -> List[int]:
    if num_splits <= 0:
        return []
    all_indices = list(range(num_splits))
    if target_count >= num_splits:
        return all_indices
    rng = np.random.default_rng(seed)
    sampled = rng.choice(num_splits, size=target_count, replace=False)
    return sorted(int(i) for i in sampled.tolist())


def _expand_crop_indices(image_indices: np.ndarray, max_length: int) -> np.ndarray:
    if image_indices.size == 0:
        return np.empty((0,), dtype=np.int64)
    expanded_indices = np.repeat(image_indices * 5, 5) + np.tile(
        np.arange(5), len(image_indices)
    )
    return expanded_indices[expanded_indices < max_length]


def _subsample_image_indices(
    image_indices: np.ndarray,
    ratio: float,
    min_count: int,
    seed: int,
) -> np.ndarray:
    image_indices = np.asarray(image_indices, dtype=np.int64)
    if image_indices.size == 0:
        return image_indices
    ratio = float(max(0.0, ratio))
    if ratio >= 1.0:
        return np.sort(image_indices)
    target_count = max(int(min_count), int(round(image_indices.size * ratio)))
    target_count = max(1, min(image_indices.size, target_count))
    if target_count >= image_indices.size:
        return np.sort(image_indices)
    rng = np.random.default_rng(seed)
    sampled = rng.choice(image_indices, size=target_count, replace=False)
    return np.sort(sampled.astype(np.int64, copy=False))


def _resolve_fast_mode_budget(args: Optional[DotMap], phase: str, default: float) -> float:
    section = _resolve_eval_section(args, phase)
    return _resolve_float_cfg(section.get("fast_mode_budget", default), default)


def _resolve_fast_mode_feature_ratio(
    args: Optional[DotMap], phase: str, default: float
) -> float:
    section = _resolve_eval_section(args, phase)
    ratio = _resolve_float_cfg(section.get("fast_mode_feature_ratio", default), default)
    return float(max(0.01, min(1.0, ratio)))


def _resolve_fast_mode_alpha_points(args: Optional[DotMap], phase: str, default: int) -> int:
    section = _resolve_eval_section(args, phase)
    return int(max(3, _resolve_int_cfg(section.get("fast_mode_alpha_points", default), default)))


def _resolve_fast_mode_alpha_min(args: Optional[DotMap], phase: str, default: float) -> float:
    section = _resolve_eval_section(args, phase)
    value = _resolve_float_cfg(section.get("fast_mode_alpha_min", default), default)
    return float(max(1e-8, value))


def _resolve_fast_mode_alpha_max(args: Optional[DotMap], phase: str, default: float) -> float:
    section = _resolve_eval_section(args, phase)
    value = _resolve_float_cfg(section.get("fast_mode_alpha_max", default), default)
    return float(max(1e-8, value))


def _resolve_fast_mode_n_jobs(args: Optional[DotMap], phase: str, default: int = 8) -> int:
    section = _resolve_eval_section(args, phase)
    return int(max(1, _resolve_int_cfg(section.get("fast_mode_n_jobs", default), default)))


def _resolve_eval_split_indices(
    dataset: Dataset,
    num_splits: int,
    phase: str,
    args: Optional[DotMap],
) -> List[int]:
    all_indices = list(range(int(max(0, num_splits))))
    if not all_indices or args is None:
        return all_indices

    section = _resolve_eval_section(args, phase)
    fast_mode = _resolve_bool_cfg(section.get("fast_mode", False), False)
    if not fast_mode:
        return all_indices

    reduce_default = phase == "val"
    reduce_eval_splits = _resolve_bool_cfg(
        section.get("fast_mode_reduce_eval_splits", reduce_default), reduce_default
    )
    if not reduce_eval_splits:
        return all_indices

    budget = _normalize_none_like(section.get("fast_mode_eval_budget", None))
    if budget is None:
        budget = _normalize_none_like(section.get("fast_mode_budget", None))
    if budget is None:
        return all_indices

    target_count = _resolve_budget_count(len(all_indices), budget, default_fraction=1.0)
    if target_count >= len(all_indices):
        return all_indices
    dataset_key = f"{dataset.__class__.__name__}:{phase}"
    seed = _resolve_seed(args, dataset_key, offset=313)
    return _sample_split_indices(len(all_indices), target_count, seed)


def _finalize_alpha_search(
    alphas: np.ndarray, srocc_all: List[List[float]]
) -> Optional[Tuple[float, int, List[float], float]]:
    if len(alphas) == 0 or len(srocc_all) == 0:
        return None

    srocc_all_median: List[float] = []
    for srocc_values in srocc_all:
        valid_values = [float(v) for v in srocc_values if v is not None and not np.isnan(v)]
        if valid_values:
            srocc_all_median.append(float(np.median(valid_values)))
        else:
            srocc_all_median.append(float("nan"))
    srocc_array = np.asarray(srocc_all_median, dtype=np.float64)
    if srocc_array.size == 0 or np.all(np.isnan(srocc_array)):
        return None
    best_alpha_idx = int(np.nanargmax(srocc_array))
    best_alpha = float(alphas[best_alpha_idx])
    return best_alpha, best_alpha_idx, srocc_all_median, float(srocc_array[best_alpha_idx])


def alpha_grid_search_budget(
    dataset: Dataset,
    features: np.ndarray,
    scores: np.ndarray,
    num_splits: int,
    split_budget: float = 0.4,
    alpha_points: int = 24,
    alpha_min: float = 1e-2,
    alpha_max: float = 1e2,
    random_seed: int = 27,
) -> Optional[Tuple[float, int, List[float], float]]:
    if num_splits <= 0:
        return None
    alpha_points = max(3, int(alpha_points))
    alpha_min = max(1e-8, float(alpha_min))
    alpha_max = max(alpha_min * 1.0001, float(alpha_max))
    alphas = np.geomspace(alpha_min, alpha_max, alpha_points, endpoint=True)
    srocc_all = [[] for _ in range(len(alphas))]

    split_count = _resolve_budget_count(num_splits, split_budget, default_fraction=0.4)
    split_indices = _sample_split_indices(num_splits, split_count, random_seed)

    for split_idx in split_indices:
        train_image_indices = np.asarray(
            dataset.get_split_indices(split=split_idx, phase="train"), dtype=np.int64
        )
        val_image_indices = np.asarray(
            dataset.get_split_indices(split=split_idx, phase="val"), dtype=np.int64
        )

        train_indices = _expand_crop_indices(train_image_indices, features.shape[0])
        val_indices = _expand_crop_indices(val_image_indices, features.shape[0])
        if train_indices.size == 0 or val_indices.size == 0:
            continue

        train_features = features[train_indices]
        train_scores = scores[train_indices]
        val_features = features[val_indices]
        val_scores = scores[val_indices][::5]

        if train_features.shape[0] < 1 or val_features.shape[0] < 1:
            continue

        for alpha_idx, alpha_value in enumerate(alphas):
            regressor = Ridge(alpha=float(alpha_value)).fit(train_features, train_scores)
            preds = regressor.predict(val_features)
            preds = np.mean(np.reshape(preds, (-1, 5)), axis=1)
            srocc_all[alpha_idx].append(float(stats.spearmanr(preds, val_scores)[0]))

    return _finalize_alpha_search(alphas, srocc_all)


def alpha_grid_search_subsample(
    dataset: Dataset,
    features: np.ndarray,
    scores: np.ndarray,
    num_splits: int,
    split_budget: float = 0.5,
    feature_ratio: float = 0.35,
    alpha_points: int = 20,
    alpha_min: float = 1e-3,
    alpha_max: float = 1e3,
    random_seed: int = 27,
) -> Optional[Tuple[float, int, List[float], float]]:
    if num_splits <= 0:
        return None
    alpha_points = max(3, int(alpha_points))
    feature_ratio = float(max(0.01, min(1.0, feature_ratio)))
    alpha_min = max(1e-8, float(alpha_min))
    alpha_max = max(alpha_min * 1.0001, float(alpha_max))
    alphas = np.geomspace(alpha_min, alpha_max, alpha_points, endpoint=True)
    srocc_all = [[] for _ in range(len(alphas))]

    split_count = _resolve_budget_count(num_splits, split_budget, default_fraction=0.5)
    split_indices = _sample_split_indices(num_splits, split_count, random_seed)

    for split_idx in split_indices:
        train_image_indices = np.asarray(
            dataset.get_split_indices(split=split_idx, phase="train"), dtype=np.int64
        )
        val_image_indices = np.asarray(
            dataset.get_split_indices(split=split_idx, phase="val"), dtype=np.int64
        )

        train_image_indices = _subsample_image_indices(
            train_image_indices,
            ratio=feature_ratio,
            min_count=64,
            seed=random_seed + split_idx * 17 + 1,
        )
        val_image_indices = _subsample_image_indices(
            val_image_indices,
            ratio=feature_ratio,
            min_count=32,
            seed=random_seed + split_idx * 17 + 2,
        )

        train_indices = _expand_crop_indices(train_image_indices, features.shape[0])
        val_indices = _expand_crop_indices(val_image_indices, features.shape[0])
        if train_indices.size == 0 or val_indices.size == 0:
            continue

        train_features = features[train_indices]
        train_scores = scores[train_indices]
        val_features = features[val_indices]
        val_scores = scores[val_indices][::5]
        if train_features.shape[0] < 1 or val_features.shape[0] < 1:
            continue

        for alpha_idx, alpha_value in enumerate(alphas):
            regressor = Ridge(alpha=float(alpha_value)).fit(train_features, train_scores)
            preds = regressor.predict(val_features)
            preds = np.mean(np.reshape(preds, (-1, 5)), axis=1)
            srocc_all[alpha_idx].append(float(stats.spearmanr(preds, val_scores)[0]))

    return _finalize_alpha_search(alphas, srocc_all)


def _resolve_alpha_cache_path(args: Optional[DotMap], phase: str) -> Path:
    section = _resolve_eval_section(args, phase)
    default_path = "results/alpha_cache/alpha_cache.json"
    raw_path = _resolve_str_cfg(section.get("fast_mode_alpha_cache_path", default_path), default_path)
    cache_path = Path(str(raw_path))
    if not cache_path.is_absolute():
        cache_path = PROJECT_ROOT / cache_path
    return cache_path


def _resolve_alpha_cache_enabled(args: Optional[DotMap], phase: str) -> bool:
    section = _resolve_eval_section(args, phase)
    default_value = phase in {"test", "cross"}
    return _resolve_bool_cfg(
        section.get("fast_mode_alpha_cache_enabled", default_value),
        default_value,
    )


def _resolve_alpha_cache_fallback_strategy(args: Optional[DotMap], phase: str) -> str:
    section = _resolve_eval_section(args, phase)
    fallback = _resolve_str_cfg(
        section.get("fast_mode_cache_fallback_strategy", "budget"), "budget"
    )
    fallback = fallback.lower()
    if fallback not in {"slow", "legacy", "budget", "subsample"}:
        fallback = "budget"
    return fallback


def _resolve_model_cache_signature(args: Optional[DotMap], phase: str) -> str:
    if args is None:
        return f"{phase}:unknown"

    explicit_signature = _resolve_str_cfg(getattr(args, "alpha_cache_signature", None), None)
    if explicit_signature:
        return explicit_signature

    eval_type = str(getattr(args, "eval_type", "scratch"))
    if eval_type != "scratch":
        return f"{phase}:unsupported_eval_type"

    experiment_name = _resolve_str_cfg(getattr(args, "experiment_name", None), None)
    if not experiment_name:
        return f"{phase}:unknown"

    checkpoint_dir = PROJECT_ROOT / "experiments" / experiment_name / "pretrain"
    if checkpoint_dir.exists():
        candidates = sorted(
            checkpoint_dir.glob("*.pth"),
            key=lambda path: path.stat().st_mtime,
            reverse=True,
        )
        preferred = [c for c in candidates if "best" in c.name]
        if not preferred:
            preferred = [c for c in candidates if "last" in c.name]
        if not preferred:
            preferred = candidates
        if preferred:
            ckpt = preferred[0]
            ckpt_stat = ckpt.stat()
            return (
                f"{experiment_name}:{ckpt.name}:"
                f"{int(ckpt_stat.st_mtime)}:{int(ckpt_stat.st_size)}"
            )
    return f"{experiment_name}:runtime"


def _resolve_dataset_signature(dataset: Dataset) -> str:
    parts = [
        dataset.__class__.__name__,
        str(len(dataset)),
        str(getattr(dataset, "mos_type", "mos")),
    ]
    images = getattr(dataset, "images", None)
    if images is not None and len(images) > 0:
        parts.append(Path(str(images[0])).name)
        parts.append(Path(str(images[-1])).name)
    return "|".join(parts)


def _load_alpha_cache(cache_path: Path) -> Dict[str, Dict[str, Any]]:
    cache_key = str(cache_path.resolve())
    cached = _ALPHA_CACHE_MEMORY.get(cache_key, None)
    if cached is not None:
        return cached

    entries: Dict[str, Dict[str, Any]] = {}
    if cache_path.exists():
        try:
            payload = json.loads(cache_path.read_text())
            if isinstance(payload, dict):
                if "entries" in payload and isinstance(payload["entries"], dict):
                    entries = payload["entries"]
                else:
                    entries = payload
        except Exception:
            entries = {}
    _ALPHA_CACHE_MEMORY[cache_key] = entries
    return entries


def _persist_alpha_cache(cache_path: Path, entries: Dict[str, Dict[str, Any]]) -> None:
    try:
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "version": _ALPHA_CACHE_FILE_VERSION,
            "updated_at": datetime.now().isoformat(),
            "entries": entries,
        }
        tmp_path = cache_path.with_suffix(cache_path.suffix + ".tmp")
        tmp_path.write_text(json.dumps(payload, indent=2, sort_keys=True))
        tmp_path.replace(cache_path)
    except Exception as exc:
        print(f"Warning: failed to persist alpha cache at {cache_path}: {exc}")


def _build_alpha_cache_key(
    dataset: Dataset,
    num_splits: int,
    strategy: str,
    settings: Dict[str, Any],
    model_signature: str,
) -> str:
    payload = {
        "dataset": _resolve_dataset_signature(dataset),
        "num_splits": int(num_splits),
        "strategy": str(strategy),
        "settings": settings,
        "model_signature": model_signature,
    }
    payload_json = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha1(payload_json.encode("utf-8")).hexdigest()


def _run_alpha_search_strategy_impl(
    strategy: str,
    dataset: Dataset,
    features: np.ndarray,
    scores: np.ndarray,
    num_splits: int,
    phase: str,
    args: Optional[DotMap] = None,
    dataset_key: Optional[str] = None,
) -> Optional[Tuple[float, int, List[float], float]]:
    strategy = (strategy or "slow").lower()
    key = dataset_key or dataset.__class__.__name__
    search_seed = _resolve_seed(args, key, offset=101)

    if strategy == "slow":
        return alpha_grid_search_1(dataset, features, scores, num_splits)
    if strategy == "legacy":
        return alpha_grid_search_11(
            dataset,
            features,
            scores,
            num_splits,
            n_jobs=_resolve_fast_mode_n_jobs(args, phase, default=8),
        )
    if strategy == "budget":
        return alpha_grid_search_budget(
            dataset=dataset,
            features=features,
            scores=scores,
            num_splits=num_splits,
            split_budget=_resolve_fast_mode_budget(args, phase, default=0.4),
            alpha_points=_resolve_fast_mode_alpha_points(args, phase, default=24),
            alpha_min=_resolve_fast_mode_alpha_min(args, phase, default=1e-2),
            alpha_max=_resolve_fast_mode_alpha_max(args, phase, default=1e2),
            random_seed=search_seed,
        )
    if strategy == "subsample":
        return alpha_grid_search_subsample(
            dataset=dataset,
            features=features,
            scores=scores,
            num_splits=num_splits,
            split_budget=_resolve_fast_mode_budget(args, phase, default=0.5),
            feature_ratio=_resolve_fast_mode_feature_ratio(args, phase, default=0.35),
            alpha_points=_resolve_fast_mode_alpha_points(args, phase, default=20),
            alpha_min=_resolve_fast_mode_alpha_min(args, phase, default=1e-3),
            alpha_max=_resolve_fast_mode_alpha_max(args, phase, default=1e3),
            random_seed=search_seed,
        )
    return alpha_grid_search_1(dataset, features, scores, num_splits)


def _resolve_alpha_with_strategy(
    dataset: Dataset,
    features: np.ndarray,
    scores: np.ndarray,
    num_splits: int,
    phase: str,
    args: Optional[DotMap] = None,
    dataset_key: Optional[str] = None,
) -> Optional[Tuple[float, int, List[float], float]]:
    strategy = _resolve_fast_mode_strategy(args, phase)
    if strategy != "cache":
        return _run_alpha_search_strategy_impl(
            strategy=strategy,
            dataset=dataset,
            features=features,
            scores=scores,
            num_splits=num_splits,
            phase=phase,
            args=args,
            dataset_key=dataset_key,
        )

    fallback_strategy = _resolve_alpha_cache_fallback_strategy(args, phase)
    if not _resolve_alpha_cache_enabled(args, phase):
        return _run_alpha_search_strategy_impl(
            strategy=fallback_strategy,
            dataset=dataset,
            features=features,
            scores=scores,
            num_splits=num_splits,
            phase=phase,
            args=args,
            dataset_key=dataset_key,
        )

    settings = {
        "fallback_strategy": fallback_strategy,
        "budget": _resolve_fast_mode_budget(args, phase, default=0.4),
        "feature_ratio": _resolve_fast_mode_feature_ratio(args, phase, default=0.35),
        "alpha_points": _resolve_fast_mode_alpha_points(args, phase, default=24),
        "alpha_min": _resolve_fast_mode_alpha_min(args, phase, default=1e-3),
        "alpha_max": _resolve_fast_mode_alpha_max(args, phase, default=1e3),
        "n_jobs": _resolve_fast_mode_n_jobs(args, phase, default=8),
        "phase": phase,
    }
    model_signature = _resolve_model_cache_signature(args, phase)
    cache_path = _resolve_alpha_cache_path(args, phase)
    cache_entries = _load_alpha_cache(cache_path)
    cache_key = _build_alpha_cache_key(
        dataset=dataset,
        num_splits=num_splits,
        strategy="cache",
        settings=settings,
        model_signature=model_signature,
    )

    cache_entry = cache_entries.get(cache_key, None)
    if isinstance(cache_entry, dict) and "alpha" in cache_entry:
        cached_alpha = float(cache_entry["alpha"])
        cached_score = float(cache_entry.get("score", float("nan")))
        dataset_label = dataset_key or dataset.__class__.__name__
        print(
            f"Alpha cache hit: dataset={dataset_label} signature={model_signature} alpha={cached_alpha:.6g}"
        )
        return cached_alpha, -1, [cached_score], cached_score

    result = _run_alpha_search_strategy_impl(
        strategy=fallback_strategy,
        dataset=dataset,
        features=features,
        scores=scores,
        num_splits=num_splits,
        phase=phase,
        args=args,
        dataset_key=dataset_key,
    )
    if result is None:
        return None

    best_alpha = float(result[0])
    best_score = float(result[3]) if result[3] is not None else float("nan")
    cache_entries[cache_key] = {
        "alpha": best_alpha,
        "score": best_score,
        "model_signature": model_signature,
        "fallback_strategy": fallback_strategy,
        "updated_at": datetime.now().isoformat(),
    }
    _persist_alpha_cache(cache_path, cache_entries)
    dataset_label = dataset_key or dataset.__class__.__name__
    print(
        f"Alpha cache miss: dataset={dataset_label} signature={model_signature} -> cached alpha={best_alpha:.6g}"
    )
    return result


def write_fr_results(
    summary_rows: List[dict],
    per_seed_rows: List[dict],
    output_dir: Path,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_path = output_dir / "fr_results.csv"
    summary_fields = [
        "dataset",
        "mode",
        "srcc",
        "plcc",
        "n_seeds",
        "aggregation",
        "split_protocol",
        "ridge_alpha",
        "ckpt",
        "timestamp",
    ]
    with open(summary_path, "w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=summary_fields)
        writer.writeheader()
        for row in summary_rows:
            writer.writerow(row)

    if per_seed_rows:
        per_seed_path = output_dir / "per_seed.csv"
        per_seed_fields = [
            "dataset",
            "seed",
            "srcc",
            "plcc",
        ]
        with open(per_seed_path, "w", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=per_seed_fields)
            writer.writeheader()
            for row in per_seed_rows:
                writer.writerow(row)


def write_fr_run_summary(
    output_dir: Path,
    args: DotMap,
    eval_type: str,
    plcc_mode: str,
    mapping_status: str,
    mapping_fn: str,
    split_protocols: dict,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    summary_path = output_dir / "run_summary.txt"
    datasets = args.test.get("datasets", [])
    lines = [
        f"timestamp={datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}",
        f"experiment_name={getattr(args, 'experiment_name', '')}",
        f"eval_type={eval_type}",
        f"plcc_mode={plcc_mode}",
        f"plcc_mapping={mapping_status} ({mapping_fn})",
        f"datasets={datasets}",
        f"num_splits={args.test.get('num_splits', '')}",
        f"aggregation=median",
        f"fast_mode={bool(getattr(args.test, 'fast_mode', False))}",
        f"grid_search={bool(getattr(args.test, 'grid_search', False))}",
        f"alpha={args.test.get('alpha', '')}",
        f"batch_size={args.test.get('batch_size', '')}",
        f"crop_size={args.test.get('crop_size', '')}",
        f"fr_iqa={resolve_fr_flag(args)}",
        f"fr_sanity_checks={resolve_fr_sanity_flag(args)}",
        f"fr_sanity_samples={resolve_fr_sanity_samples(args)}",
    ]
    lines.append("split_protocols:")
    for dataset_name, protocol in split_protocols.items():
        lines.append(f"- {dataset_name}: {protocol}")
    with open(summary_path, "w") as handle:
        handle.write("\n".join(lines) + "\n")


def log_fr_results_to_wandb(
    logger: Optional[Run],
    summary_rows: List[dict],
    per_seed_rows: List[dict],
    output_dir: Path,
) -> None:
    if not logger:
        return

    if summary_rows:
        summary_fields = [
            "dataset",
            "mode",
            "srcc",
            "plcc",
            "n_seeds",
            "aggregation",
            "split_protocol",
            "ridge_alpha",
            "ckpt",
            "timestamp",
        ]
        summary_table = wandb.Table(
            columns=summary_fields,
            data=[[row.get(field) for field in summary_fields] for row in summary_rows],
        )
        logger.log({"fr_results": summary_table})

        srcc_values = []
        plcc_values = []
        for row in summary_rows:
            dataset_key = sanitize_identifier(str(row.get("dataset", "")))
            logger.summary[f"fr_srocc_{dataset_key}"] = row.get("srcc")
            logger.summary[f"fr_plcc_{dataset_key}"] = row.get("plcc")
            logger.summary[f"fr_alpha_{dataset_key}"] = row.get("ridge_alpha")
            logger.summary[f"fr_n_seeds_{dataset_key}"] = row.get("n_seeds")
            logger.summary[f"fr_split_protocol_{dataset_key}"] = row.get(
                "split_protocol"
            )
            logger.summary[f"fr_ckpt_{dataset_key}"] = row.get("ckpt")
            logger.summary[f"fr_timestamp_{dataset_key}"] = row.get("timestamp")
            if row.get("srcc") is not None:
                srcc_values.append(row.get("srcc"))
            if row.get("plcc") is not None:
                plcc_values.append(row.get("plcc"))

        logger.summary["fr_srocc_avg"] = safe_mean(srcc_values)
        logger.summary["fr_plcc_avg"] = safe_mean(plcc_values)
        logger.summary["fr_results_path"] = str(output_dir)

    if per_seed_rows:
        per_seed_fields = [
            "dataset",
            "seed",
            "srcc",
            "plcc",
        ]
        per_seed_table = wandb.Table(
            columns=per_seed_fields,
            data=[[row.get(field) for field in per_seed_fields] for row in per_seed_rows],
        )
        logger.log({"fr_per_seed": per_seed_table})


class FRSubsetDataset(Dataset):
    def __init__(self, dataset: Dataset, indices: List[int]):
        self.dataset = dataset
        self.indices = list(indices)
        self._index_map = {orig: idx for idx, orig in enumerate(self.indices)}

        self.is_synthetic = getattr(dataset, "is_synthetic", False)
        self.mos_type = getattr(dataset, "mos_type", "mos")
        self.mos_range = getattr(dataset, "mos_range", None)
        self.images = np.asarray(dataset.images)[self.indices]
        self.ref_images = np.asarray(dataset.ref_images)[self.indices]
        self.mos = np.asarray(dataset.mos)[self.indices]
        if self.is_synthetic:
            self.distortion_types = np.asarray(dataset.distortion_types)[self.indices]
            self.distortion_groups = np.asarray(dataset.distortion_groups)[self.indices]
            self.distortion_levels = np.asarray(dataset.distortion_levels)[self.indices]

    def __len__(self) -> int:
        return len(self.indices)

    def __getitem__(self, idx: int) -> dict:
        return self.dataset[self.indices[idx]]

    def get_split_indices(self, split: int, phase: str) -> np.ndarray:
        orig_indices = self.dataset.get_split_indices(split, phase)
        subset_indices = [
            self._index_map[orig] for orig in orig_indices if orig in self._index_map
        ]
        return np.array(subset_indices, dtype=int)


def _extract_ref_id_from_paths(dataset_name: str, dist_path: Path, ref_path: Path):
    dataset_name = dataset_name.lower()
    if dataset_name == "live":
        return ref_path.stem
    if dataset_name == "csiq":
        match = re.search(r"(\w+)\.(\w+)\.(\d+)", dist_path.name)
        return match.group(1) if match else None
    if dataset_name == "tid2013":
        match = re.search(r"(i\d+)_\d+_\d+", dist_path.stem, re.IGNORECASE)
        return match.group(1) if match else None
    if dataset_name == "kadid10k":
        return dist_path.stem.split("_")[0]
    return None


def _ref_id_matches(dataset_name: str, dist_path: Path, ref_path: Path) -> bool:
    ref_id = _extract_ref_id_from_paths(dataset_name, dist_path, ref_path)
    if ref_id is None:
        return True
    ref_stem = ref_path.stem.lower()
    return ref_id.lower() in ref_stem


def run_fr_sanity_checks(
    args: DotMap,
    model: nn.Module,
    device: torch.device,
    eval_type: str,
) -> None:
    datasets = args.test.get("datasets", [])
    sanity_samples = resolve_fr_sanity_samples(args)
    mapping_status, mapping_fn = describe_plcc_mapping(args)

    print("FR sanity checks enabled.")
    print(f"PLCC logistic mapping: {mapping_status}, function={mapping_fn}")

    for d in datasets:
        dataset, dataset_num_splits, dataset_display_name = prepare_dataset(
            d,
            data_base_path=args.data_base_path,
            crop_size=args.test.crop_size,
            default_num_splits=args.test.num_splits,
            fr_iqa=True,
        )

        split_summary = describe_split_protocol(
            dataset, dataset_num_splits, aggregation="median"
        )
        print(f"Splits: {split_summary}")

        label_type = getattr(dataset, "mos_type", "mos").upper()
        print(
            f"Dataset={dataset_display_name} label_type={label_type} "
            "sign_flip_applied=False"
        )

        sample_count = min(sanity_samples, len(dataset))
        print(f"Pairing check ({dataset_display_name}) first {sample_count} samples:")
        ref_map = {}
        for idx in range(sample_count):
            dist_path = Path(dataset.images[idx])
            ref_path = Path(dataset.ref_images[idx])
            score = float(dataset.mos[idx])
            if not ref_path.exists():
                raise FileNotFoundError(f"Missing ref image: {ref_path}")
            print(f"  {idx}: dist={dist_path} ref={ref_path} score={score:.4f}")

            ref_id = _extract_ref_id_from_paths(d, dist_path, ref_path)
            if ref_id is not None:
                ref_map.setdefault(ref_id, set()).add(str(ref_path))
            if not _ref_id_matches(d, dist_path, ref_path):
                raise ValueError(
                    f"Ref ID mismatch for {dataset_display_name}: "
                    f"dist={dist_path} ref={ref_path}"
                )

        if d.lower() == "live" and ref_map:
            for ref_id, paths in ref_map.items():
                if len(paths) > 1:
                    raise ValueError(
                        f"LIVE ref_id {ref_id} maps to multiple references: {paths}"
                    )

        sample_a = dataset[0]
        sample_b = dataset[0]
        img_dist = sample_a["img"]
        img_ref = sample_a["ref_img"]
        img_dist_ds = sample_a["img_ds"]
        img_ref_ds = sample_a["ref_img_ds"]
        print(
            f"Transform check shapes: dist={tuple(img_dist.shape)} "
            f"ref={tuple(img_ref.shape)} dist_ds={tuple(img_dist_ds.shape)} "
            f"ref_ds={tuple(img_ref_ds.shape)}"
        )
        if not torch.allclose(img_dist, sample_b["img"]) or not torch.allclose(
            img_ref, sample_b["ref_img"]
        ):
            raise ValueError(
                "Detected non-deterministic transforms in FR mode; "
                "paired transforms are required."
            )
        print("Transforms: deterministic center/corner crops (paired transform not required).")

        subset_indices = list(range(sample_count))
        subset = FRSubsetDataset(dataset, subset_indices)
        sanity_loader = DataLoader(
            subset,
            batch_size=min(args.test.batch_size, 8),
            shuffle=False,
            num_workers=0,
            pin_memory=resolve_eval_pin_memory(args=args, default=False),
        )
        batch = next(iter(sanity_loader))
        amp_context = (
            torch.autocast(device_type="cuda")
            if device.type == "cuda" and torch.cuda.is_available()
            else nullcontext()
        )
        with amp_context, torch.no_grad():
            dist = rearrange(batch["img"].to(device), "b n c h w -> (b n) c h w")
            dist_ds = rearrange(
                batch["img_ds"].to(device), "b n c h w -> (b n) c h w"
            )
            ref = rearrange(batch["ref_img"].to(device), "b n c h w -> (b n) c h w")
            ref_ds = rearrange(
                batch["ref_img_ds"].to(device), "b n c h w -> (b n) c h w"
            )
            if eval_type == "scratch":
                _, g_dist = model(dist)
                _, g_dist_ds = model(dist_ds)
                h_dist = torch.cat((g_dist, g_dist_ds), dim=1)
                _, g_ref = model(ref)
                _, g_ref_ds = model(ref_ds)
                h_ref = torch.cat((g_ref, g_ref_ds), dim=1)
            else:
                _, h_dist = model(dist, dist_ds, return_embedding=True)
                _, h_ref = model(ref, ref_ds, return_embedding=True)

            u = torch.abs(h_ref - h_dist)
            u_min = float(u.min().item())
            u_mean = float(u.mean().item())
            u_max = float(u.max().item())
            print(
                f"Feature sanity: h_ref={tuple(h_ref.shape)} "
                f"h_dist={tuple(h_dist.shape)} u={tuple(u.shape)} "
                f"u_min={u_min:.6f} u_mean={u_mean:.6f} u_max={u_max:.6f}"
            )
            if u_min < -1e-8:
                raise ValueError("u has negative entries beyond tolerance.")

            if eval_type == "scratch":
                _, g_ref1 = model(ref)
                _, g_ref1_ds = model(ref_ds)
                h_ref1 = torch.cat((g_ref1, g_ref1_ds), dim=1)
                _, g_ref2 = model(ref)
                _, g_ref2_ds = model(ref_ds)
                h_ref2 = torch.cat((g_ref2, g_ref2_ds), dim=1)
            else:
                raise ValueError(f"Eval type {eval_type} not supported")
            u_same = torch.abs(h_ref1 - h_ref2)
            print(f"Identical pair sanity: max|u|={float(u_same.max().item()):.6f}")

        sanity_splits = min(dataset_num_splits, 1)
        srocc_dataset, plcc_dataset, _, best_alpha, _ = compute_fr_metrics(
            model=model,
            dataset=subset,
            num_splits=sanity_splits,
            phase="test",
            alpha=args.test.alpha,
            grid_search=False,
            batch_size=min(args.test.batch_size, 8),
            num_workers=0,
            device=device,
            eval_type=eval_type,
            args=args,
        )
        srocc_values = srocc_dataset.get("global", [])
        if srocc_values:
            srocc_val = srocc_values[0]
            sign_ok = srocc_val >= 0
            print(
                f"Sanity SRCC sign check: value={srocc_val:.4f} "
                f"label_type={label_type} sign_ok={sign_ok}"
            )
        print(f"Sanity ridge alpha (subset): {best_alpha}")


def run_fr_evaluation(
    args: DotMap,
    model: nn.Module,
    logger: Run,
    device: torch.device,
    eval_type: str,
) -> None:
    datasets = args.test.get("datasets", [])
    if not datasets:
        raise ValueError("FR-IQA requires at least one dataset.")

    invalid = [d for d in datasets if d.lower() not in synthetic_datasets]
    if invalid:
        raise ValueError(
            "FR-IQA only supports LIVE, CSIQ, TID2013, and KADID-10K. "
            f"Unsupported datasets: {', '.join(invalid)}"
        )

    run_timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    checkpoint_label = resolve_checkpoint_label(args, eval_type)
    summary_rows = []
    per_seed_rows = []
    use_logistic = resolve_plcc_logistic(args)
    plcc_mode = "plcc_logistic" if use_logistic else "plcc_raw"
    mapping_status, mapping_fn = describe_plcc_mapping(args)
    print(f"PLCC logistic mapping: {mapping_status}, function={mapping_fn}")
    split_protocols = {}
    if logger:
        logger.summary["fr_plcc_mode"] = plcc_mode
        logger.summary["fr_plcc_mapping"] = mapping_fn

    for d in datasets:
        dataset, dataset_num_splits, dataset_display_name = prepare_dataset(
            d,
            data_base_path=args.data_base_path,
            crop_size=args.test.crop_size,
            default_num_splits=args.test.num_splits,
            fr_iqa=True,
        )

        split_protocol = describe_split_protocol(
            dataset, dataset_num_splits, aggregation="median"
        )
        print(f"Splits: {split_protocol}")
        split_protocols[dataset_display_name] = split_protocol

        srocc_dataset, plcc_dataset, _, best_alpha, _ = compute_fr_metrics(
            model=model,
            dataset=dataset,
            num_splits=dataset_num_splits,
            phase="test",
            alpha=args.test.alpha,
            grid_search=args.test.grid_search,
            batch_size=args.test.batch_size,
            num_workers=args.test.num_workers,
            device=device,
            eval_type=eval_type,
            args=args,
        )

        srocc_values = srocc_dataset.get("global", [])
        plcc_values = plcc_dataset.get("global", [])
        srocc_median = (
            float(np.median(srocc_values)) if len(srocc_values) > 0 else float("nan")
        )
        plcc_median = (
            float(np.median(plcc_values)) if len(plcc_values) > 0 else float("nan")
        )

        print(
            f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - "
            f"Dataset={dataset_display_name} Mode=FR "
            f"SRCC={format_float(srocc_median)} PLCC={format_float(plcc_median)}"
        )

        summary_rows.append(
            {
                "dataset": dataset_display_name,
                "mode": "FR",
                "srcc": srocc_median,
                "plcc": plcc_median,
                "n_seeds": dataset_num_splits,
                "aggregation": "median",
                "split_protocol": split_protocol,
                "ridge_alpha": best_alpha,
                "ckpt": checkpoint_label,
                "timestamp": run_timestamp,
            }
        )

        for split_idx, srocc_value in enumerate(srocc_values):
            plcc_value = (
                plcc_values[split_idx]
                if split_idx < len(plcc_values)
                else float("nan")
            )
            per_seed_rows.append(
                {
                    "dataset": dataset_display_name,
                    "seed": split_idx,
                    "srcc": srocc_value,
                    "plcc": plcc_value,
                }
            )

    output_dir = PROJECT_ROOT / "results" / "fr_iqa" / f"{args.experiment_name}_{plcc_mode}"
    log_fr_results_to_wandb(logger, summary_rows, per_seed_rows, output_dir)
    write_fr_results(summary_rows, per_seed_rows, output_dir)
    write_fr_run_summary(
        output_dir=output_dir,
        args=args,
        eval_type=eval_type,
        plcc_mode=plcc_mode,
        mapping_status=mapping_status,
        mapping_fn=mapping_fn,
        split_protocols=split_protocols,
    )
    print(f"FR results saved to {output_dir}")

def test(args: DotMap, model: nn.Module, logger: Run, device: torch.device) -> None:
    """
    Test pretrained model on the test datasets. Performs a grid search over the validation splits to find the best
    alpha value for the regression for each dataset. Saves a CSV file with the results and a pickle file with the
    regressor for each dataset.

    Args:
        args (dotmap.DotMap): test arguments
        model (torch.nn.Module): model to test
        logger (wandb.wandb_run.Run): wandb logger
        device (torch.device): device to use for testing
    """
    checkpoint_base_path = PROJECT_ROOT / "experiments"
    checkpoint_path = checkpoint_base_path / args.experiment_name
    regressor_path = checkpoint_path / "regressors"
    regressor_path.mkdir(parents=True, exist_ok=True)

    eval_type = args.get("eval_type", "scratch")

    model.eval()

    eval_log_step: Optional[int] = None
    try:
        eval_log_step = int(getattr(args, "global_step", 0))
    except (TypeError, ValueError):
        eval_log_step = 0
    if eval_log_step <= 0 and hasattr(args, "training"):
        try:
            eval_log_step = int(getattr(args.training, "max_steps", 0))
        except (TypeError, ValueError):
            eval_log_step = 0
    if eval_log_step <= 0:
        eval_log_step = None

    cross_cfg = args.test.get("cross_eval", DotMap())
    cross_enabled = bool(cross_cfg.get("enabled", False))
    cross_include_standard = _resolve_bool_cfg(
        cross_cfg.get("include_standard", True), True
    )
    fr_iqa = resolve_fr_flag(args)
    fr_sanity = resolve_fr_sanity_flag(args)

    if fr_iqa:
        if getattr(args.test, "fast_mode", False):
            print("FR-IQA: overriding test.fast_mode=False (fast mode disabled).")
            args.test.fast_mode = False
        if cross_enabled:
            raise ValueError("Cross-dataset evaluation is not supported in FR-IQA mode.")
        if fr_sanity:
            run_fr_sanity_checks(args, model, device, eval_type)
        run_fr_evaluation(args, model, logger, device, eval_type)
        return

    if cross_enabled:
        train_datasets = cross_cfg.get("train_datasets", [])
        test_datasets = cross_cfg.get("test_datasets", [])
        if not train_datasets:
            raise ValueError("Cross-dataset evaluation requires at least one train dataset")
        if not test_datasets:
            raise ValueError("Cross-dataset evaluation requires at least one test dataset")
        cross_args = copy.deepcopy(args)
        cross_override_keys = [
            "fast_mode",
            "fast_mode_strategy",
            "fast_mode_budget",
            "fast_mode_feature_ratio",
            "fast_mode_alpha_points",
            "fast_mode_alpha_min",
            "fast_mode_alpha_max",
            "fast_mode_reduce_eval_splits",
            "fast_mode_eval_budget",
            "fast_mode_alpha_cache_enabled",
            "fast_mode_cache_fallback_strategy",
            "fast_mode_alpha_cache_path",
            "fast_mode_n_jobs",
        ]
        for key in cross_override_keys:
            value = _normalize_none_like(cross_cfg.get(key, None))
            if value is not None:
                cross_args.test.cross_eval[key] = value
            elif hasattr(args.test, "get"):
                fallback = _normalize_none_like(args.test.get(key, None))
                if fallback is not None:
                    cross_args.test.cross_eval[key] = fallback
        cross_grid_search = _resolve_bool_cfg(
            cross_cfg.get("grid_search", args.test.grid_search),
            bool(args.test.grid_search),
        )
        cross_alpha = _resolve_float_cfg(
            cross_cfg.get("alpha", args.test.alpha), float(args.test.alpha)
        )
        cross_args.test.cross_eval.grid_search = cross_grid_search
        cross_args.test.cross_eval.alpha = cross_alpha
        print(
            "Cross-eval strategy config: "
            f"grid_search={cross_grid_search} "
            f"fast_mode={_resolve_bool_cfg(cross_args.test.cross_eval.get('fast_mode', args.test.fast_mode), bool(args.test.fast_mode))} "
            f"strategy={_resolve_fast_mode_strategy(cross_args, 'cross')}"
        )
        (
            cross_srocc_all,
            cross_plcc_all,
            cross_regressors,
            cross_alphas,
            cross_best_worst_results_all,
            cross_vicreg_metrics_all,
            cross_result_metadata,
        ) = get_cross_results(
            model=model,
            data_base_path=args.data_base_path,
            train_datasets=train_datasets,
            test_datasets=test_datasets,
            num_splits=args.test.num_splits,
            phase="test",
            alpha=cross_alpha,
            grid_search=cross_grid_search,
            crop_size=args.test.crop_size,
            batch_size=args.test.batch_size,
            num_workers=args.test.num_workers,
            device=device,
            eval_type=eval_type,
            args=cross_args,
            logger=logger,
        )
        if cross_include_standard:
            (
                srocc_all,
                plcc_all,
                regressors,
                alphas,
                best_worst_results_all,
                vicreg_metrics_all,
                result_metadata,
            ) = get_results(
                model=model,
                data_base_path=args.data_base_path,
                datasets=args.test.datasets,
                num_splits=args.test.num_splits,
                phase="test",
                alpha=args.test.alpha,
                grid_search=args.test.grid_search,
                crop_size=args.test.crop_size,
                batch_size=args.test.batch_size,
                num_workers=args.test.num_workers,
                device=device,
                eval_type=eval_type,
                num_logging_steps=eval_log_step,
                args=args,
                logger=logger,
            )
            srocc_all.update(cross_srocc_all)
            plcc_all.update(cross_plcc_all)
            regressors.update(cross_regressors)
            alphas.update(cross_alphas)
            best_worst_results_all.update(cross_best_worst_results_all)
            vicreg_metrics_all.update(cross_vicreg_metrics_all)
            result_metadata.update(cross_result_metadata)
        else:
            (
                srocc_all,
                plcc_all,
                regressors,
                alphas,
                best_worst_results_all,
                vicreg_metrics_all,
                result_metadata,
            ) = (
                cross_srocc_all,
                cross_plcc_all,
                cross_regressors,
                cross_alphas,
                cross_best_worst_results_all,
                cross_vicreg_metrics_all,
                cross_result_metadata,
            )
    else:
        (
            srocc_all,
            plcc_all,
            regressors,
            alphas,
            best_worst_results_all,
            vicreg_metrics_all,
            result_metadata,
        ) = get_results(
            model=model,
            data_base_path=args.data_base_path,
            datasets=args.test.datasets,
            num_splits=args.test.num_splits,
            phase="test",
            alpha=args.test.alpha,
            grid_search=args.test.grid_search,
            crop_size=args.test.crop_size,
            batch_size=args.test.batch_size,
            num_workers=args.test.num_workers,
            device=device,
            eval_type=eval_type,
            num_logging_steps=eval_log_step,
            args=args,
            logger=logger,
        )

    srocc_all_median = {}
    plcc_all_median = {}
    for key, value in srocc_all.items():
        srocc_values = value.get("global", [])
        srocc_all_median[key] = (
            float(np.median(srocc_values)) if len(srocc_values) > 0 else float("nan")
        )
    for key, value in plcc_all.items():
        plcc_values = value.get("global", [])
        plcc_all_median[key] = (
            float(np.median(plcc_values)) if len(plcc_values) > 0 else float("nan")
        )

    synthetic_srocc_values = [
        srocc_all_median[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in synthetic_datasets
        and not np.isnan(srocc_all_median[key])
    ]
    synthetic_plcc_values = [
        plcc_all_median[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in synthetic_datasets
        and not np.isnan(plcc_all_median[key])
    ]
    authentic_srocc_values = [
        srocc_all_median[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in authentic_datasets
        and not np.isnan(srocc_all_median[key])
    ]
    authentic_plcc_values = [
        plcc_all_median[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in authentic_datasets
        and not np.isnan(plcc_all_median[key])
    ]

    srocc_synthetic_avg = safe_mean(synthetic_srocc_values)
    plcc_synthetic_avg = safe_mean(synthetic_plcc_values)
    srocc_authentic_avg = safe_mean(authentic_srocc_values)
    plcc_authentic_avg = safe_mean(authentic_plcc_values)

    within_keys = [
        key
        for key, meta in result_metadata.items()
        if meta["train_dataset"] == meta["test_dataset"]
    ]
    cross_keys = [
        key
        for key, meta in result_metadata.items()
        if meta["train_dataset"] != meta["test_dataset"]
    ]
    has_cross_pairs = len(cross_keys) > 0

    within_srocc_values = [
        srocc_all_median[key]
        for key in within_keys
        if not np.isnan(srocc_all_median[key])
    ]
    within_plcc_values = [
        plcc_all_median[key]
        for key in within_keys
        if not np.isnan(plcc_all_median[key])
    ]
    cross_srocc_values = [
        srocc_all_median[key]
        for key in cross_keys
        if not np.isnan(srocc_all_median[key])
    ]
    cross_plcc_values = [
        plcc_all_median[key]
        for key in cross_keys
        if not np.isnan(plcc_all_median[key])
    ]

    srocc_all_best = {}
    plcc_all_best = {}
    for key, value in srocc_all.items():
        vals = [
            v for v in value.get("global", []) if v is not None and not np.isnan(v)
        ]
        srocc_all_best[key] = safe_max(vals) if vals else float("nan")
    for key, value in plcc_all.items():
        vals = [
            v for v in value.get("global", []) if v is not None and not np.isnan(v)
        ]
        plcc_all_best[key] = safe_max(vals) if vals else float("nan")

    synthetic_srocc_best_values = [
        srocc_all_best[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in synthetic_datasets
        and not np.isnan(srocc_all_best[key])
    ]
    synthetic_plcc_best_values = [
        plcc_all_best[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in synthetic_datasets
        and not np.isnan(plcc_all_best[key])
    ]
    authentic_srocc_best_values = [
        srocc_all_best[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in authentic_datasets
        and not np.isnan(srocc_all_best[key])
    ]
    authentic_plcc_best_values = [
        plcc_all_best[key]
        for key, meta in result_metadata.items()
        if meta["test_dataset"] in authentic_datasets
        and not np.isnan(plcc_all_best[key])
    ]

    within_srocc_best_values = [
        srocc_all_best[key] for key in within_keys if not np.isnan(srocc_all_best[key])
    ]
    within_plcc_best_values = [
        plcc_all_best[key] for key in within_keys if not np.isnan(plcc_all_best[key])
    ]
    cross_srocc_best_values = [
        srocc_all_best[key] for key in cross_keys if not np.isnan(srocc_all_best[key])
    ]
    cross_plcc_best_values = [
        plcc_all_best[key] for key in cross_keys if not np.isnan(plcc_all_best[key])
    ]

    srocc_synthetic_best_avg = safe_mean(synthetic_srocc_best_values)
    plcc_synthetic_best_avg = safe_mean(synthetic_plcc_best_values)
    srocc_authentic_best_avg = safe_mean(authentic_srocc_best_values)
    plcc_authentic_best_avg = safe_mean(authentic_plcc_best_values)

    srocc_within_avg = safe_mean(within_srocc_values)
    plcc_within_avg = safe_mean(within_plcc_values)
    srocc_cross_avg = safe_mean(cross_srocc_values)
    plcc_cross_avg = safe_mean(cross_plcc_values)

    srocc_within_best_avg = safe_mean(within_srocc_best_values)
    plcc_within_best_avg = safe_mean(within_plcc_best_values)
    srocc_cross_best_avg = safe_mean(cross_srocc_best_values)
    plcc_cross_best_avg = safe_mean(cross_plcc_best_values)

    srocc_values_all = [val for val in srocc_all_median.values() if not np.isnan(val)]
    plcc_values_all = [val for val in plcc_all_median.values() if not np.isnan(val)]
    srocc_avg = safe_mean(srocc_values_all)
    plcc_avg = safe_mean(plcc_values_all)

    validation_datasets_cfg = (
        list(getattr(args.validation, "datasets", []))
        if hasattr(args, "validation")
        else list(args.test.datasets)
    )
    validation_dataset_keys = {str(ds).lower() for ds in validation_datasets_cfg}
    val_within_keys = [
        key
        for key in within_keys
        if result_metadata[key]["test_dataset"] in validation_dataset_keys
    ]
    if not val_within_keys:
        val_within_keys = list(within_keys)

    val_srocc_per_dataset = {}
    val_plcc_per_dataset = {}
    for key in val_within_keys:
        val_srocc = float(result_metadata[key].get("val_srocc_median", float("nan")))
        val_plcc = float(result_metadata[key].get("val_plcc_median", float("nan")))
        if np.isnan(val_srocc):
            val_srocc = float(srocc_all_median[key])
        if np.isnan(val_plcc):
            val_plcc = float(plcc_all_median[key])
        val_srocc_per_dataset[key] = val_srocc
        val_plcc_per_dataset[key] = val_plcc

    val_synthetic_srocc_values = [
        val_srocc_per_dataset[key]
        for key in val_within_keys
        if result_metadata[key]["test_dataset"] in synthetic_datasets
        and not np.isnan(val_srocc_per_dataset[key])
    ]
    val_synthetic_plcc_values = [
        val_plcc_per_dataset[key]
        for key in val_within_keys
        if result_metadata[key]["test_dataset"] in synthetic_datasets
        and not np.isnan(val_plcc_per_dataset[key])
    ]
    val_authentic_srocc_values = [
        val_srocc_per_dataset[key]
        for key in val_within_keys
        if result_metadata[key]["test_dataset"] in authentic_datasets
        and not np.isnan(val_srocc_per_dataset[key])
    ]
    val_authentic_plcc_values = [
        val_plcc_per_dataset[key]
        for key in val_within_keys
        if result_metadata[key]["test_dataset"] in authentic_datasets
        and not np.isnan(val_plcc_per_dataset[key])
    ]

    val_srocc_synthetic_avg = safe_mean(val_synthetic_srocc_values)
    val_plcc_synthetic_avg = safe_mean(val_synthetic_plcc_values)
    val_srocc_authentic_avg = safe_mean(val_authentic_srocc_values)
    val_plcc_authentic_avg = safe_mean(val_authentic_plcc_values)
    val_srocc_avg = safe_mean(
        [val for val in val_srocc_per_dataset.values() if not np.isnan(val)]
    )
    val_plcc_avg = safe_mean(
        [val for val in val_plcc_per_dataset.values() if not np.isnan(val)]
    )
    val_within_srocc_avg = val_srocc_avg
    val_within_plcc_avg = val_plcc_avg
    if has_cross_pairs:
        val_cross_srocc_avg = srocc_cross_avg
        val_cross_plcc_avg = plcc_cross_avg
    else:
        val_cross_srocc_avg = float("nan")
        val_cross_plcc_avg = float("nan")
    if np.isfinite(val_cross_srocc_avg):
        val_srocc_blend = float(0.5 * val_srocc_avg + 0.5 * val_cross_srocc_avg)
    else:
        val_srocc_blend = float(val_srocc_avg)

    srocc_best_values_all = [
        val for val in srocc_all_best.values() if not np.isnan(val)
    ]
    plcc_best_values_all = [
        val for val in plcc_all_best.values() if not np.isnan(val)
    ]
    srocc_best_avg = safe_mean(srocc_best_values_all)
    plcc_best_avg = safe_mean(plcc_best_values_all)

    label_width = 30 if has_cross_pairs else 15
    header_label = "Train->Test" if has_cross_pairs else "Dataset"
    print(f"{header_label:<{label_width}} {'Alpha':<15} {'SROCC':<15} {'PLCC':<15}")
    for key in result_metadata.keys():
        meta = result_metadata[key]
        alpha_str = format_alpha(alphas[key])
        srocc_str = format_float(srocc_all_median[key])
        plcc_str = format_float(plcc_all_median[key])
        print(
            f"{meta['display_name']:<{label_width}} {alpha_str:<15} {srocc_str:<15} {plcc_str:<15}"
        )
    print(
        f"{'Synthetic avg':<{label_width}} {'':<15} {format_float(srocc_synthetic_avg):<15} {format_float(plcc_synthetic_avg):<15}"
    )
    print(
        f"{'Authentic avg':<{label_width}} {'':<15} {format_float(srocc_authentic_avg):<15} {format_float(plcc_authentic_avg):<15}"
    )
    if has_cross_pairs:
        print(
            f"{'Within avg':<{label_width}} {'':<15} {format_float(srocc_within_avg):<15} {format_float(plcc_within_avg):<15}"
        )
        print(
            f"{'Cross avg':<{label_width}} {'':<15} {format_float(srocc_cross_avg):<15} {format_float(plcc_cross_avg):<15}"
        )
    print(
        f"{'Global avg':<{label_width}} {'':<15} {format_float(srocc_avg):<15} {format_float(plcc_avg):<15}"
    )
    print(
        f"{'Val avg (from test)':<{label_width}} {'':<15} {format_float(val_srocc_avg):<15} {format_float(val_plcc_avg):<15}"
    )
    if has_cross_pairs:
        print(
            f"{'Val cross avg (proxy)':<{label_width}} {'':<15} {format_float(val_cross_srocc_avg):<15} {format_float(val_cross_plcc_avg):<15}"
        )
    print(
        f"{'Val blend':<{label_width}} {'':<15} {format_float(val_srocc_blend):<15} {'':<15}"
    )

    print("")
    best_header_label = (
        "Train->Test (best)" if has_cross_pairs else "Dataset (best)"
    )
    print(f"{best_header_label:<{label_width}} {'Alpha':<15} {'SROCC':<15} {'PLCC':<15}")
    for key in result_metadata.keys():
        meta = result_metadata[key]
        alpha_str = format_alpha(alphas[key])
        srocc_best_str = format_float(srocc_all_best[key])
        plcc_best_str = format_float(plcc_all_best[key])
        print(
            f"{meta['display_name']:<{label_width}} {alpha_str:<15} {srocc_best_str:<15} {plcc_best_str:<15}"
        )
    print(
        f"{'Synthetic avg':<{label_width}} {'':<15} {format_float(srocc_synthetic_best_avg):<15} {format_float(plcc_synthetic_best_avg):<15}"
    )
    print(
        f"{'Authentic avg':<{label_width}} {'':<15} {format_float(srocc_authentic_best_avg):<15} {format_float(plcc_authentic_best_avg):<15}"
    )
    if has_cross_pairs:
        print(
            f"{'Within avg':<{label_width}} {'':<15} {format_float(srocc_within_best_avg):<15} {format_float(plcc_within_best_avg):<15}"
        )
        print(
            f"{'Cross avg':<{label_width}} {'':<15} {format_float(srocc_cross_best_avg):<15} {format_float(plcc_cross_best_avg):<15}"
        )
    print(
        f"{'Global avg':<{label_width}} {'':<15} {format_float(srocc_best_avg):<15} {format_float(plcc_best_avg):<15}"
    )

    workbook = openpyxl.Workbook()

    median_sheet = workbook.create_sheet("Median", 0)
    median_sheet.append(["Dataset", "Alpha", "SROCC", "PLCC"])
    for key in result_metadata.keys():
        median_sheet.append(
            [
                result_metadata[key]["display_name"],
                alphas[key],
                srocc_all_median[key],
                plcc_all_median[key],
            ]
        )
    median_sheet.append(["Synthetic avg", "", srocc_synthetic_avg, plcc_synthetic_avg])
    median_sheet.append(["Authentic avg", "", srocc_authentic_avg, plcc_authentic_avg])
    if has_cross_pairs:
        median_sheet.append(
            ["Within avg (train=test)", "", srocc_within_avg, plcc_within_avg]
        )
        median_sheet.append(
            ["Cross avg (train!=test)", "", srocc_cross_avg, plcc_cross_avg]
        )
    median_sheet.append(["Global avg", "", srocc_avg, plcc_avg])

    best_sheet = workbook.create_sheet("Best", 1)
    best_sheet.append(["Dataset", "Alpha", "SROCC", "PLCC"])
    for key in result_metadata.keys():
        best_sheet.append(
            [
                result_metadata[key]["display_name"],
                alphas[key],
                srocc_all_best[key],
                plcc_all_best[key],
            ]
        )
    best_sheet.append(["Synthetic avg", "", srocc_synthetic_best_avg, plcc_synthetic_best_avg])
    best_sheet.append(["Authentic avg", "", srocc_authentic_best_avg, plcc_authentic_best_avg])
    if has_cross_pairs:
        best_sheet.append([
            "Within avg (train=test)",
            "",
            srocc_within_best_avg,
            plcc_within_best_avg,
        ])
        best_sheet.append([
            "Cross avg (train!=test)",
            "",
            srocc_cross_best_avg,
            plcc_cross_best_avg,
        ])
    best_sheet.append(["Global avg", "", srocc_best_avg, plcc_best_avg])

    used_sheet_names = {"Median", "Best"}
    for key in result_metadata.keys():
        srocc_results = srocc_all[key]
        plcc_results = plcc_all[key]
        base_sheet_name = result_metadata[key]["sheet_name"]
        sheet_name = base_sheet_name
        if sheet_name in used_sheet_names:
            idx = 1
            while True:
                suffix = f"_{idx}"
                trimmed_base = base_sheet_name[: max(0, 31 - len(suffix))]
                candidate = f"{trimmed_base}{suffix}" if trimmed_base else suffix
                candidate = sanitize_sheet_name(candidate)
                if candidate not in used_sheet_names:
                    sheet_name = candidate
                    break
                idx += 1
        used_sheet_names.add(sheet_name)
        sheet = workbook.create_sheet(sheet_name)
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        sheet.cell(row=1, column=1).value = "Split"
        sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        for i in range(1, len(srocc_results) + 1):
            sheet.merge_cells(
                start_row=1, start_column=(i * 2), end_row=1, end_column=(i * 2) + 1
            )
            sheet.cell(row=1, column=(i * 2)).value = list(srocc_results.keys())[i - 1]
            sheet.cell(row=1, column=(i * 2)).alignment = Alignment(horizontal="center")
            sheet.cell(row=2, column=(i * 2)).value = "SROCC"
            sheet.cell(row=2, column=(i * 2) + 1).value = "PLCC"

        for i in range(len(srocc_results["global"])):
            row = [i]
            for dist_type in srocc_results.keys():
                if i >= len(srocc_results[dist_type]):
                    continue
                row += [srocc_results[dist_type][i]]
                row += [plcc_results[dist_type][i]]
            sheet.append(row)

    workbook.active = 0

    workbook.remove(workbook["Sheet"])

    workbook.save(checkpoint_path / "results.xlsx")

    if logger:
        for key in result_metadata.keys():
            logger_suffix = result_metadata[key]["logger_suffix"]
            logger.summary[f"test_srocc_{logger_suffix}"] = srocc_all_median[key]
            logger.summary[f"test_plcc_{logger_suffix}"] = plcc_all_median[key]
            logger.summary[f"test_alpha_{logger_suffix}"] = alphas[key]
            logger.summary[f"test_srocc_best_{logger_suffix}"] = srocc_all_best[key]
            logger.summary[f"test_plcc_best_{logger_suffix}"] = plcc_all_best[key]
        for key in val_within_keys:
            logger_suffix = result_metadata[key]["logger_suffix"]
            logger.summary[f"val_srocc_{logger_suffix}"] = val_srocc_per_dataset[key]
            logger.summary[f"val_plcc_{logger_suffix}"] = val_plcc_per_dataset[key]
        logger.summary[f"test_srocc_synthetic_avg"] = srocc_synthetic_avg
        logger.summary[f"test_plcc_synthetic_avg"] = plcc_synthetic_avg
        logger.summary[f"test_srocc_authentic_avg"] = srocc_authentic_avg
        logger.summary[f"test_plcc_authentic_avg"] = plcc_authentic_avg
        if has_cross_pairs:
            logger.summary["test_srocc_within_avg"] = srocc_within_avg
            logger.summary["test_plcc_within_avg"] = plcc_within_avg
            logger.summary["test_srocc_cross_avg"] = srocc_cross_avg
            logger.summary["test_plcc_cross_avg"] = plcc_cross_avg
        logger.summary[f"test_srocc_avg"] = srocc_avg
        logger.summary[f"test_plcc_avg"] = plcc_avg
        logger.summary["val_srocc_synthetic_avg"] = val_srocc_synthetic_avg
        logger.summary["val_plcc_synthetic_avg"] = val_plcc_synthetic_avg
        logger.summary["val_srocc_authentic_avg"] = val_srocc_authentic_avg
        logger.summary["val_plcc_authentic_avg"] = val_plcc_authentic_avg
        logger.summary["val_srocc_avg"] = val_srocc_avg
        logger.summary["val_plcc_avg"] = val_plcc_avg
        logger.summary["val_within_srocc_avg"] = val_within_srocc_avg
        logger.summary["val_within_plcc_avg"] = val_within_plcc_avg
        logger.summary["val_cross_srocc_avg"] = val_cross_srocc_avg
        logger.summary["val_cross_plcc_avg"] = val_cross_plcc_avg
        logger.summary["val_srocc_blend"] = val_srocc_blend

        aggregate_history = {
            "test_srocc_synthetic_avg": srocc_synthetic_avg,
            "test_plcc_synthetic_avg": plcc_synthetic_avg,
            "test_srocc_authentic_avg": srocc_authentic_avg,
            "test_plcc_authentic_avg": plcc_authentic_avg,
            "test_srocc_avg": srocc_avg,
            "test_plcc_avg": plcc_avg,
            "val_srocc_synthetic_avg": val_srocc_synthetic_avg,
            "val_plcc_synthetic_avg": val_plcc_synthetic_avg,
            "val_srocc_authentic_avg": val_srocc_authentic_avg,
            "val_plcc_authentic_avg": val_plcc_authentic_avg,
            "val_srocc_avg": val_srocc_avg,
            "val_plcc_avg": val_plcc_avg,
            "val_within_srocc_avg": val_within_srocc_avg,
            "val_within_plcc_avg": val_within_plcc_avg,
            "val_cross_srocc_avg": val_cross_srocc_avg,
            "val_cross_plcc_avg": val_cross_plcc_avg,
            "val_srocc_blend": val_srocc_blend,
        }
        if has_cross_pairs:
            aggregate_history["test_srocc_within_avg"] = srocc_within_avg
            aggregate_history["test_plcc_within_avg"] = plcc_within_avg
            aggregate_history["test_srocc_cross_avg"] = srocc_cross_avg
            aggregate_history["test_plcc_cross_avg"] = plcc_cross_avg

        backfill_payload = dict(aggregate_history)
        backfill_payload["test_srocc_best_avg"] = srocc_best_avg
        backfill_payload["test_plcc_best_avg"] = plcc_best_avg
        _dump_final_eval_metrics(checkpoint_path, logger, backfill_payload)

        logger.log(aggregate_history)

        logger.summary[f"test_srocc_synthetic_best_avg"] = srocc_synthetic_best_avg
        logger.summary[f"test_plcc_synthetic_best_avg"] = plcc_synthetic_best_avg
        logger.summary[f"test_srocc_authentic_best_avg"] = srocc_authentic_best_avg
        logger.summary[f"test_plcc_authentic_best_avg"] = plcc_authentic_best_avg
        if has_cross_pairs:
            logger.summary["test_srocc_within_best_avg"] = srocc_within_best_avg
            logger.summary["test_plcc_within_best_avg"] = plcc_within_best_avg
            logger.summary["test_srocc_cross_best_avg"] = srocc_cross_best_avg
            logger.summary["test_plcc_cross_best_avg"] = plcc_cross_best_avg
        logger.summary[f"test_srocc_best_avg"] = srocc_best_avg
        logger.summary[f"test_plcc_best_avg"] = plcc_best_avg

        log_best_worst_tables = _resolve_bool_cfg(
            getattr(args.test, "log_best_worst_tables", True), True
        )
        if log_best_worst_tables:
            for result_key in best_worst_results_all.keys():
                meta = result_metadata[result_key]
                logger_suffix = meta["logger_suffix"]
                for kind in best_worst_results_all[
                    result_key
                ].keys():  # key is either "best" or "worst"
                    column_names = ["Image", "Predicted", "GT", "Difference", "Path"]
                    table_data = []
                    for i in range(
                        len(best_worst_results_all[result_key][kind]["images"])
                    ):
                        img_path = best_worst_results_all[result_key][kind]["images"][i]
                        img = (
                            wandb.Image(Image.open(img_path))
                            if args.logging.wandb.imgs
                            else None
                        )
                        gt = best_worst_results_all[result_key][kind]["gts"][i]
                        pred = best_worst_results_all[result_key][kind]["preds"][i]
                        diff = pred - gt
                        table_data.append([img, pred, gt, diff, str(img_path)])
                    logger.log(
                        {
                            f"test_{logger_suffix}_{kind}_results": wandb.Table(
                                data=table_data, columns=column_names
                            )
                        }
                    )

    for key, regressor in regressors.items():
        meta = result_metadata[key]
        filename = (
            f"{meta['file_stem']}_srocc_{format_float(srocc_all_median[key])}"
            f"_plcc_{format_float(plcc_all_median[key])}.pkl"
        )
        with open(regressor_path / filename, "wb") as f:
            pickle.dump(regressor, f)


def get_results(
    model: nn.Module,
    data_base_path: Path,
    datasets: List[str],
    num_splits: int,
    phase: str,
    alpha: float,
    grid_search: bool,
    crop_size: int,
    batch_size: int,
    num_workers: int,
    device: torch.device,
    eval_type: str = "scratch",
    eval_dtype: torch.dtype = torch.float64,
    num_logging_steps=None,
    logger=None,
    args=None,
) -> Tuple[dict, dict, dict, dict, dict]:
    """
    Get the results for the given model and datasets. Depending on the phase parameter, can be used both for validation
    and test. If phase == 'test' and grid_search == True, performs a grid search over the validation splits to find the best
    alpha value for the regression for each dataset. The results related to synthetic datasets contain also the results
    for each distortion type.

    Args:
        model (torch.nn.Module): model to test
        data_base_path (pathlib.Path): base path of the datasets
        datasets (list): list of datasets
        num_splits (int): number of splits
        phase (str): phase of the datasets. Must be in ['val', 'test']
        alpha (float): alpha value to use for regression. During test, if None, performs a grid search
        grid_search (bool): whether to perform a grid search over the validation splits to find the best alpha value for the regression
        crop_size (int): crop size
        batch_size (int): batch size
        num_workers (int): number of workers for the dataloaders
        device (torch.device): device to use for testing
        eval_type (str): Evaluation mode. This release supports only `scratch`.

    Returns:
        srocc_all (dict): dictionary containing the SROCC results
        plcc_all (dict): dictionary containing the PLCC results
        regressors (dict): dictionary containing the regressors
        alphas (dict): dictionary containing the alpha values used for the regression
        best_worst_results_all (dict): dictionary containing the best and worst results
        vicreg_metrics_all (dict): VICReg-inspired metrics
        result_metadata (dict): metadata for formatting and logging
    """
    srocc_all = {}
    plcc_all = {}
    regressors = {}
    alphas = {}
    best_worst_results_all = {}
    vicreg_metrics_all = {}
    result_metadata = {}
    logged_vicreg_metrics = set()
    per_dataset_timing: Dict[str, float] = {}

    assert phase in ["val", "test"], "Phase must be in ['val', 'test']"

    print(f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - Starting {phase} phase")
    phase_start_time = time.perf_counter()
    for d in datasets:
        dataset_start_time = time.perf_counter()
        dataset, dataset_num_splits, dataset_display_name = prepare_dataset(
            d,
            data_base_path=data_base_path,
            crop_size=crop_size,
            default_num_splits=num_splits,
        )

        (
            srocc_dataset,
            plcc_dataset,
            regressor,
            alpha,
            best_worst_results,
            vicreg_metrics_dataset,
            shared_phase_metrics_dataset,
        ) = compute_metrics(
            model,
            dataset,
            dataset_num_splits,
            phase,
            alpha,
            grid_search,
            batch_size,
            num_workers,
            device,
            eval_type,
            eval_dtype,
            args,
        )
        srocc_all[d] = srocc_dataset
        plcc_all[d] = plcc_dataset
        regressors[d] = regressor
        alphas[d] = alpha
        best_worst_results_all[d] = best_worst_results
        vicreg_metrics_all[d] = vicreg_metrics_dataset
        result_metadata[d] = build_result_metadata(
            train_key=d.lower(),
            test_key=d.lower(),
            train_display=dataset_display_name,
            test_display=dataset_display_name,
        )
        if shared_phase_metrics_dataset:
            if "val_srocc_median" in shared_phase_metrics_dataset:
                result_metadata[d]["val_srocc_median"] = float(
                    shared_phase_metrics_dataset.get("val_srocc_median", float("nan"))
                )
            if "val_plcc_median" in shared_phase_metrics_dataset:
                result_metadata[d]["val_plcc_median"] = float(
                    shared_phase_metrics_dataset.get("val_plcc_median", float("nan"))
                )
            if "test_srocc_median" in shared_phase_metrics_dataset:
                result_metadata[d]["test_srocc_median"] = float(
                    shared_phase_metrics_dataset.get("test_srocc_median", float("nan"))
                )
            if "test_plcc_median" in shared_phase_metrics_dataset:
                result_metadata[d]["test_plcc_median"] = float(
                    shared_phase_metrics_dataset.get("test_plcc_median", float("nan"))
                )
        print(
            f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - {dataset_display_name}:"
            f" SRCC: {np.median(srocc_dataset['global']):.3f} - PLCC: {np.median(plcc_dataset['global']):.3f}"
        )
        print(f"vicreg_metrics_dataset: {vicreg_metrics_dataset}")
        if logger is not None:
            if d not in logged_vicreg_metrics:
                vicreg_payload = {
                    f"{key}_{d}": vicreg_metrics_dataset[key]
                    for key in vicreg_metrics_dataset.keys()
                }
                # Keep default step semantics to ensure eval metrics are retained.
                logger.log(vicreg_payload)
                logged_vicreg_metrics.add(d)

        dataset_elapsed = time.perf_counter() - dataset_start_time
        per_dataset_timing[d] = float(dataset_elapsed)
        print(
            f"TIMING phase={phase} dataset={d.lower()} seconds={dataset_elapsed:.3f}"
        )

        gc.collect()
        if device.type == "cuda" and torch.cuda.is_available():
            torch.cuda.empty_cache()

    phase_elapsed = time.perf_counter() - phase_start_time
    print(f"TIMING_TOTAL phase={phase} seconds={phase_elapsed:.3f}")
    if args is not None:
        try:
            args._timing_stats = DotMap(
                {
                    "phase": phase,
                    "total_seconds": float(phase_elapsed),
                    "per_dataset_seconds": per_dataset_timing,
                }
            )
        except Exception:
            pass

    return (
        srocc_all,
        plcc_all,
        regressors,
        alphas,
        best_worst_results_all,
        vicreg_metrics_all,
        result_metadata,
    )


def get_cross_results(
    model: nn.Module,
    data_base_path: Path,
    train_datasets: List[str],
    test_datasets: List[str],
    num_splits: int,
    phase: str,
    alpha: float,
    grid_search: bool,
    crop_size: int,
    batch_size: int,
    num_workers: int,
    device: torch.device,
    eval_type: str = "scratch",
    args=None,
    logger: Optional[Run] = None,
):
    assert phase in ["val", "test"], "Phase must be in ['val', 'test']"

    print(
        f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - Starting cross-dataset {phase} phase"
    )

    canonical_train = [ds.lower() for ds in train_datasets]
    canonical_test = [ds.lower() for ds in test_datasets]

    dataset_cache = {}
    feature_cache = {}
    use_logistic = resolve_plcc_logistic(args)

    unique_keys = []
    for key in canonical_train + canonical_test:
        if key not in unique_keys:
            unique_keys.append(key)

    for key in unique_keys:
        dataset, dataset_num_splits, dataset_display_name = prepare_dataset(
            key,
            data_base_path=data_base_path,
            crop_size=crop_size,
            default_num_splits=num_splits,
        )
        dataloader = DataLoader(
            dataset,
            batch_size=batch_size,
            shuffle=False,
            num_workers=num_workers,
            pin_memory=resolve_eval_pin_memory(args=args, default=False),
        )
        reps, embs, scores = get_features_scores(
            model, dataloader, device, eval_type, dtype=torch.float64
        )
        dataset_cache[key] = {
            "dataset": dataset,
            "num_splits": dataset_num_splits,
            "display_name": dataset_display_name,
        }
        feature_cache[key] = {"reps": reps, "embs": embs, "scores": scores}

    alpha_cache = {}

    for train_key in dict.fromkeys(canonical_train):
        if grid_search:
            search_result = _resolve_alpha_with_strategy(
                dataset=dataset_cache[train_key]["dataset"],
                features=feature_cache[train_key]["reps"],
                scores=feature_cache[train_key]["scores"],
                num_splits=dataset_cache[train_key]["num_splits"],
                phase="cross",
                args=args,
                dataset_key=train_key,
            )
            best_alpha = search_result[0] if search_result is not None else alpha
        else:
            best_alpha = alpha
        alpha_cache[train_key] = best_alpha

    srocc_all = {}
    plcc_all = {}
    regressors = {}
    alphas = {}
    best_worst_results_all = {}
    vicreg_metrics_all = {}
    result_metadata = {}

    for train_key in canonical_train:
        train_info = dataset_cache[train_key]
        train_features = feature_cache[train_key]["reps"]
        train_scores = feature_cache[train_key]["scores"]
        train_num_splits = train_info["num_splits"]
        best_alpha = alpha_cache[train_key]

        for test_key in canonical_test:
            test_info = dataset_cache[test_key]
            test_features = feature_cache[test_key]["reps"]
            test_scores = feature_cache[test_key]["scores"]
            test_embeddings = feature_cache[test_key]["embs"]
            (
                srocc_dataset,
                plcc_dataset,
                regressor,
                _,
                best_worst_results,
                vicreg_metrics_dataset,
            ) = compute_cross_dataset_metrics(
                train_dataset=train_info["dataset"],
                test_dataset=test_info["dataset"],
                train_features=train_features,
                train_scores=train_scores,
                test_features=test_features,
                test_scores=test_scores,
                train_num_splits=train_num_splits,
                test_num_splits=test_info["num_splits"],
                phase=phase,
                best_alpha=best_alpha,
                test_embeddings=test_embeddings,
                device=device,
                plcc_logistic=use_logistic,
                compute_vicreg=False,
            )

            result_key = f"{train_key}->{test_key}"
            srocc_all[result_key] = srocc_dataset
            plcc_all[result_key] = plcc_dataset
            regressors[result_key] = regressor
            alphas[result_key] = best_alpha
            best_worst_results_all[result_key] = best_worst_results
            vicreg_metrics_all[result_key] = vicreg_metrics_dataset
            result_metadata[result_key] = build_result_metadata(
                train_key=train_key,
                test_key=test_key,
                train_display=train_info["display_name"],
                test_display=test_info["display_name"],
            )
            result_metadata[result_key].update(
                {
                    "train_mos_type": getattr(train_info["dataset"], "mos_type", "mos"),
                    "test_mos_type": getattr(test_info["dataset"], "mos_type", "mos"),
                    "train_mos_range": getattr(
                        train_info["dataset"], "mos_range", None
                    ),
                    "test_mos_range": getattr(
                        test_info["dataset"], "mos_range", None
                    ),
                }
            )

            srocc_values = srocc_dataset.get("global", [])
            plcc_values = plcc_dataset.get("global", [])
            srocc_median = (
                float(np.median(srocc_values)) if len(srocc_values) > 0 else float("nan")
            )
            plcc_median = (
                float(np.median(plcc_values)) if len(plcc_values) > 0 else float("nan")
            )
            print(
                f"{datetime.now().strftime('%d/%m/%Y %H:%M:%S')} - "
                f"{train_info['display_name']} -> {test_info['display_name']}:"
                f" SRCC: {srocc_median:.3f} - PLCC: {plcc_median:.3f}"
            )

    return (
        srocc_all,
        plcc_all,
        regressors,
        alphas,
        best_worst_results_all,
        vicreg_metrics_all,
        result_metadata,
    )


def compute_cross_dataset_metrics(
    train_dataset: Dataset,
    test_dataset: Dataset,
    train_features: np.ndarray,
    train_scores: np.ndarray,
    test_features: np.ndarray,
    test_scores: np.ndarray,
    train_num_splits: int,
    test_num_splits: int,
    phase: str,
    best_alpha: float,
    test_embeddings: np.ndarray,
    device: torch.device,
    plcc_logistic: bool = False,
    compute_vicreg: bool = True,
) -> Tuple[dict, dict, Ridge, float, dict, dict]:
    train_mos_type = getattr(train_dataset, "mos_type", "mos")
    train_mos_range = getattr(train_dataset, "mos_range", None)
    test_mos_type = getattr(test_dataset, "mos_type", "mos")
    test_mos_range = getattr(test_dataset, "mos_range", None)

    srocc_dataset = {"global": []}
    plcc_dataset = {"global": []}
    best_worst_results = {"best": {}, "worst": {}}

    unique_dist_types = []
    if test_dataset.is_synthetic:
        unique_dist_types = sorted(set(test_dataset.distortion_types))
        for dist_type in unique_dist_types:
            srocc_dataset[dist_type] = []
            plcc_dataset[dist_type] = []

    test_scores_per_crop = test_scores
    if test_scores_per_crop.size == 0:
        empty_srocc = {"global": [float("nan")]}
        empty_plcc = {"global": [float("nan")]}
        return (
            empty_srocc,
            empty_plcc,
            None,
            best_alpha,
            {"best": {}, "worst": {}},
            {},
        )

    test_scores_per_image = test_scores_per_crop[::5]
    images_array = np.asarray(test_dataset.images)

    effective_splits = train_num_splits if train_num_splits > 0 else 1

    for split_idx in range(effective_splits):
        train_indices = train_dataset.get_split_indices(split_idx, "train")
        if len(train_indices) == 0:
            continue

        train_indices_expanded = np.repeat(train_indices * 5, 5) + np.tile(
            np.arange(5), len(train_indices)
        )
        train_indices_expanded = train_indices_expanded[
            train_indices_expanded < train_features.shape[0]
        ]
        if train_indices_expanded.size == 0:
            continue

        train_features_split = train_features[train_indices_expanded]
        train_scores_split = train_scores[train_indices_expanded]

        regressor_split = Ridge(alpha=best_alpha).fit(
            train_features_split, train_scores_split
        )

        preds = regressor_split.predict(test_features)
        preds = np.mean(np.reshape(preds, (-1, 5)), axis=1)
        preds_aligned = map_scores_between_datasets(
            preds,
            source_type=train_mos_type,
            source_range=train_mos_range,
            target_type=test_mos_type,
            target_range=test_mos_range,
        )

        srocc_dataset["global"].append(
            stats.spearmanr(preds_aligned, test_scores_per_image)[0]
        )
        plcc_preds_aligned = preds_aligned
        if plcc_logistic:
            plcc_preds_aligned = apply_plcc_logistic_mapping(
                preds_aligned, test_scores_per_image
            )
        plcc_dataset["global"].append(
            stats.pearsonr(plcc_preds_aligned, test_scores_per_image)[0]
        )

        if unique_dist_types:
            for dist_type in unique_dist_types:
                dist_indices = np.where(test_dataset.distortion_types == dist_type)[0]
                if dist_indices.size < 2:
                    continue
                srocc_dataset[dist_type].append(
                    stats.spearmanr(
                        preds_aligned[dist_indices], test_scores_per_image[dist_indices]
                    )[0]
                )
                plcc_dataset[dist_type].append(
                    stats.pearsonr(
                        plcc_preds_aligned[dist_indices],
                        test_scores_per_image[dist_indices],
                    )[0]
                )

        if split_idx == 0:
            diff = np.abs(preds_aligned - test_scores_per_image)
            sorted_diff_indices = np.argsort(diff)
            best_indices = sorted_diff_indices[:16]
            worst_indices = sorted_diff_indices[-16:][::-1]
            best_images = (
                images_array[best_indices].tolist() if images_array.size else []
            )
            worst_images = (
                images_array[worst_indices].tolist() if images_array.size else []
            )
            best_worst_results = {
                "best": {
                    "images": best_images,
                    "gts": test_scores_per_image[best_indices],
                    "preds": preds_aligned[best_indices],
                },
                "worst": {
                    "images": worst_images,
                    "gts": test_scores_per_image[worst_indices],
                    "preds": preds_aligned[worst_indices],
                },
            }

    if len(srocc_dataset["global"]) == 0:
        srocc_dataset["global"].append(float("nan"))
        plcc_dataset["global"].append(float("nan"))
    for dist_type in unique_dist_types:
        if len(srocc_dataset[dist_type]) == 0:
            srocc_dataset[dist_type].append(float("nan"))
            plcc_dataset[dist_type].append(float("nan"))

    regressor_full = None
    if train_features.shape[0] > 0:
        regressor_full = Ridge(alpha=best_alpha).fit(train_features, train_scores)

    if compute_vicreg:
        vicreg_metrics = compute_vicreg_metrics(
            reps=test_features, embs=test_embeddings, device=device
        )
    else:
        vicreg_metrics = {}

    return (
        srocc_dataset,
        plcc_dataset,
        regressor_full,
        best_alpha,
        best_worst_results,
        vicreg_metrics,
    )


def evaluate_ridge_metrics(
    dataset: Dataset,
    features: np.ndarray,
    scores: np.ndarray,
    num_splits: int,
    phase: str,
    alpha: float,
    grid_search: bool,
    args=None,
) -> Tuple[dict, dict, Ridge, float, dict]:
    srocc_dataset = {"global": []}
    plcc_dataset = {"global": []}
    best_worst_results = {}
    dist_types = None
    if dataset.is_synthetic:
        dist_types = set(dataset.distortion_types)
        for dist_type in dist_types:
            srocc_dataset[dist_type] = []
            plcc_dataset[dist_type] = []

    if grid_search:
        search_result = _resolve_alpha_with_strategy(
            dataset=dataset,
            features=features,
            scores=scores,
            num_splits=num_splits,
            phase=phase,
            args=args,
            dataset_key=dataset.__class__.__name__,
        )
        best_alpha = search_result[0] if search_result is not None else alpha
    else:
        best_alpha = alpha

    split_indices = _resolve_eval_split_indices(
        dataset=dataset,
        num_splits=num_splits,
        phase=phase,
        args=args,
    )
    if not split_indices:
        split_indices = list(range(max(0, num_splits)))

    for split_pos, split_idx in enumerate(split_indices):
        train_image_indices = np.asarray(
            dataset.get_split_indices(split=split_idx, phase="train"), dtype=np.int64
        )
        eval_image_indices = np.asarray(
            dataset.get_split_indices(split=split_idx, phase=phase), dtype=np.int64
        )

        dist_indices = None
        if dataset.is_synthetic:
            dist_indices = {
                dist_type: np.where(
                    dataset.distortion_types[eval_image_indices] == dist_type
                )[0]
                for dist_type in dist_types
            }

        train_indices = _expand_crop_indices(train_image_indices, features.shape[0])
        eval_indices = _expand_crop_indices(eval_image_indices, features.shape[0])

        train_features = features[train_indices]
        train_scores = scores[train_indices]

        if train_features.shape[0] < 1:
            continue

        regressor = Ridge(alpha=best_alpha).fit(train_features, train_scores)

        test_features = features[eval_indices]
        test_scores = scores[eval_indices]
        test_scores = test_scores[
            ::5
        ]  # Scores are repeated for each crop, so we only keep the first one
        orig_test_indices = eval_indices[::5] // 5  # Get original indices

        if test_features.shape[0] < 1:
            continue
        preds = regressor.predict(test_features)
        preds = np.mean(
            np.reshape(preds, (-1, 5)), 1
        )  # Average the predictions of the 5 crops of the same image

        srocc_dataset["global"].append(stats.spearmanr(preds, test_scores)[0])
        plcc_preds = preds
        if resolve_plcc_logistic(args):
            plcc_preds = apply_plcc_logistic_mapping(preds, test_scores)
        plcc_dataset["global"].append(stats.pearsonr(plcc_preds, test_scores)[0])

        if dataset.is_synthetic:
            for dist_type in dist_types:
                indices = dist_indices[dist_type]
                indices = indices[indices < preds.shape[0]]
                if indices.shape[0] < 2:
                    continue

                srocc_dataset[dist_type].append(
                    stats.spearmanr(
                        preds[indices],
                        test_scores[indices],
                    )[0]
                )
                plcc_dataset[dist_type].append(
                    stats.pearsonr(
                        plcc_preds[indices],
                        test_scores[indices],
                    )[0]
                )

        if split_pos == 0:
            diff = np.abs(preds - test_scores)
            sorted_diff_indices = np.argsort(diff)
            best_indices = sorted_diff_indices[:16]
            worst_indices = sorted_diff_indices[-16:][::-1]
            best_worst_results["best"] = {
                "images": dataset.images[orig_test_indices[best_indices]],
                "gts": test_scores[best_indices],
                "preds": preds[best_indices],
            }
            best_worst_results["worst"] = {
                "images": dataset.images[orig_test_indices[worst_indices]],
                "gts": test_scores[worst_indices],
                "preds": preds[worst_indices],
            }

    if len(srocc_dataset["global"]) == 0:
        srocc_dataset["global"].append(float("nan"))
        plcc_dataset["global"].append(float("nan"))
    if dataset.is_synthetic:
        for dist_type in dist_types:
            if len(srocc_dataset[dist_type]) == 0:
                srocc_dataset[dist_type].append(float("nan"))
                plcc_dataset[dist_type].append(float("nan"))

    regressor_full = None
    if features.shape[0] > 0:
        regressor_full = Ridge(alpha=best_alpha).fit(features, scores)

    return (
        srocc_dataset,
        plcc_dataset,
        regressor_full,
        best_alpha,
        best_worst_results,
    )


def evaluate_joint_val_test_with_shared_regressor(
    dataset: Dataset,
    features: np.ndarray,
    scores: np.ndarray,
    num_splits: int,
    alpha: float,
    grid_search: bool,
    args=None,
) -> Dict[str, Any]:
    """
    Fit one regressor per split (train split) and evaluate both val/test from that same fit.
    This is useful to check whether the two-phase eval can be merged without changing numbers.
    """

    if grid_search:
        search_result = _resolve_alpha_with_strategy(
            dataset=dataset,
            features=features,
            scores=scores,
            num_splits=num_splits,
            phase="test",
            args=args,
            dataset_key=dataset.__class__.__name__,
        )
        best_alpha = search_result[0] if search_result is not None else alpha
    else:
        best_alpha = alpha

    split_indices = _resolve_eval_split_indices(
        dataset=dataset,
        num_splits=num_splits,
        phase="test",
        args=args,
    )
    if not split_indices:
        split_indices = list(range(max(0, num_splits)))

    out = {
        "alpha": float(best_alpha),
        "val": {"srocc": [], "plcc": []},
        "test": {"srocc": [], "plcc": []},
    }

    use_logistic = resolve_plcc_logistic(args)
    for split_idx in split_indices:
        train_image_indices = np.asarray(
            dataset.get_split_indices(split=split_idx, phase="train"), dtype=np.int64
        )
        train_indices = _expand_crop_indices(train_image_indices, features.shape[0])
        if train_indices.size < 1:
            continue

        regressor = Ridge(alpha=best_alpha).fit(
            features[train_indices], scores[train_indices]
        )

        for eval_phase in ("val", "test"):
            eval_image_indices = np.asarray(
                dataset.get_split_indices(split=split_idx, phase=eval_phase),
                dtype=np.int64,
            )
            eval_indices = _expand_crop_indices(eval_image_indices, features.shape[0])
            if eval_indices.size < 1:
                continue

            eval_features = features[eval_indices]
            eval_scores = scores[eval_indices][::5]
            if eval_features.shape[0] < 1:
                continue

            preds = regressor.predict(eval_features)
            preds = np.mean(np.reshape(preds, (-1, 5)), axis=1)
            out["val" if eval_phase == "val" else "test"]["srocc"].append(
                stats.spearmanr(preds, eval_scores)[0]
            )
            plcc_preds = preds
            if use_logistic:
                plcc_preds = apply_plcc_logistic_mapping(preds, eval_scores)
            out["val" if eval_phase == "val" else "test"]["plcc"].append(
                stats.pearsonr(plcc_preds, eval_scores)[0]
            )

    def _safe_median(values):
        if len(values) == 0:
            return float("nan")
        return float(np.median(values))

    out["val"]["srocc_median"] = _safe_median(out["val"]["srocc"])
    out["val"]["plcc_median"] = _safe_median(out["val"]["plcc"])
    out["test"]["srocc_median"] = _safe_median(out["test"]["srocc"])
    out["test"]["plcc_median"] = _safe_median(out["test"]["plcc"])
    return out


def compute_metrics(
    model: nn.Module,
    dataset: Dataset,
    num_splits: int,
    phase: str,
    alpha: float,
    grid_search: bool,
    batch_size: int,
    num_workers: int,
    device: torch.device,
    eval_type: str = "scratch",
    eval_dtype: torch.dtype = torch.float64,
    args=None,
) -> Tuple[dict, dict, Ridge, float, dict, dict, dict]:
    """
    Compute the metrics for the given model and dataset. If phase == 'test' and grid_search == True, performs a grid search
    over the validation splits to find the best alpha value for the regression.

    Args:
        model (torch.nn.Module): model to test
        dataset (torch.utils.data.Dataset): dataset to test on
        num_splits (int): number of splits
        phase (str): phase of the datasets. Must be in ['val', 'test']
        alpha (float): alpha value to use for regression. During test, if None, performs a grid search
        grid_search (bool): whether to perform a grid search over the validation splits to find the best alpha value for the regression
        batch_size (int): batch size
        num_workers (int): number of workers for the dataloaders
        device (torch.device): device to use for testing
        eval_type (str): Evaluation mode. This release supports only `scratch`.

    Returns:
        srocc_dataset (dict): dictionary containing the SROCC results for the dataset
        plcc_dataset (dict): dictionary containing the PLCC results for the dataset
        regressor (Ridge): Ridge regressor
        alpha (float): alpha value used for the regression
        best_worst_results (dict): dictionary containing the best and worst results
        vicreg_metrics (dict): VICReg-inspired metrics
        shared_phase_metrics (dict): opposite-phase metrics computed from the same
            train-fit (val from test phase, or test from val phase)
    """
    dataloader = DataLoader(
        dataset,
        batch_size=batch_size,
        shuffle=False,
        num_workers=num_workers,
        pin_memory=resolve_eval_pin_memory(args=args, default=False),
    )

    reps, embs, scores = get_features_scores(
        model, dataloader, device, eval_type, dtype=eval_dtype
    )

    (
        srocc_dataset,
        plcc_dataset,
        regressor,
        best_alpha,
        best_worst_results,
    ) = evaluate_ridge_metrics(
        dataset=dataset,
        features=reps,
        scores=scores,
        num_splits=num_splits,
        phase=phase,
        alpha=alpha,
        grid_search=grid_search,
        args=args,
    )

    report_val_from_test = True
    report_test_from_val = False
    if args is not None:
        if hasattr(args, "test") and args.test is not None:
            report_val_from_test = _resolve_bool_cfg(
                args.test.get("report_val_from_test_phase", True), True
            )
        if hasattr(args, "validation") and args.validation is not None:
            report_test_from_val = _resolve_bool_cfg(
                args.validation.get("report_test_from_val_phase", False), False
            )

    shared_phase_metrics = {}
    needs_joint_eval = (phase == "test" and report_val_from_test) or (
        phase == "val" and report_test_from_val
    )
    if needs_joint_eval:
        joint_eval = evaluate_joint_val_test_with_shared_regressor(
            dataset=dataset,
            features=reps,
            scores=scores,
            num_splits=num_splits,
            alpha=alpha,
            grid_search=grid_search,
            args=args,
        )
        if phase == "test":
            shared_phase_metrics["val_srocc_median"] = float(
                joint_eval["val"]["srocc_median"]
            )
            shared_phase_metrics["val_plcc_median"] = float(
                joint_eval["val"]["plcc_median"]
            )
        elif phase == "val":
            shared_phase_metrics["test_srocc_median"] = float(
                joint_eval["test"]["srocc_median"]
            )
            shared_phase_metrics["test_plcc_median"] = float(
                joint_eval["test"]["plcc_median"]
            )

    # VICReg-inspired metrics
    vicreg_metrics = compute_vicreg_metrics(reps=reps, embs=embs, device=device)

    return (
        srocc_dataset,
        plcc_dataset,
        regressor,
        best_alpha,
        best_worst_results,
        vicreg_metrics,
        shared_phase_metrics,
    )


def compute_fr_features(
    model: nn.Module,
    dataloader: DataLoader,
    device: torch.device,
    eval_type: str = "scratch",
    dtype=torch.float32,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Get FR features u = |h_ref - h_dist| and scores for the given dataloader.
    """
    num_samples = len(dataloader.dataset) * 5
    features = None
    scores = None
    idx = 0
    for batch in dataloader:
        if "ref_img" not in batch or "ref_img_ds" not in batch:
            raise ValueError(
                "FR mode requires dataset batches to include 'ref_img' and 'ref_img_ds'."
            )
        img_dist = batch["img"].to(device)
        img_dist_ds = batch["img_ds"].to(device)
        img_ref = batch["ref_img"].to(device)
        img_ref_ds = batch["ref_img_ds"].to(device)
        mos = batch["mos"]

        img_dist = rearrange(img_dist, "b n c h w -> (b n) c h w")
        img_dist_ds = rearrange(img_dist_ds, "b n c h w -> (b n) c h w")
        img_ref = rearrange(img_ref, "b n c h w -> (b n) c h w")
        img_ref_ds = rearrange(img_ref_ds, "b n c h w -> (b n) c h w")
        mos = mos.repeat_interleave(5)

        amp_context = (
            torch.autocast(device_type="cuda")
            if device.type == "cuda" and torch.cuda.is_available()
            else nullcontext()
        )
        with amp_context, torch.no_grad():
            if eval_type == "scratch":
                _, g_dist = model(img_dist)
                _, g_dist_ds = model(img_dist_ds)
                h_dist = torch.cat((g_dist, g_dist_ds), dim=1)

                _, g_ref = model(img_ref)
                _, g_ref_ds = model(img_ref_ds)
                h_ref = torch.cat((g_ref, g_ref_ds), dim=1)
            else:
                raise ValueError(f"Eval type {eval_type} not supported")

        u = torch.abs(h_ref - h_dist)
        batch_size = u.size(0)
        if features is None:
            feat_dim = u.size(1)
            features = torch.empty((num_samples, feat_dim), dtype=dtype, device=device)
            scores = torch.empty((num_samples,), dtype=dtype, device="cpu")
        features[idx : idx + batch_size].copy_(u)
        scores[idx : idx + batch_size].copy_(mos)
        idx += batch_size

    if features is None or scores is None:
        return np.empty((0, 0)), np.empty((0,))
    return features[:idx].cpu().numpy(), scores[:idx].numpy()


def compute_fr_metrics(
    model: nn.Module,
    dataset: Dataset,
    num_splits: int,
    phase: str,
    alpha: float,
    grid_search: bool,
    batch_size: int,
    num_workers: int,
    device: torch.device,
    eval_type: str = "scratch",
    args=None,
) -> Tuple[dict, dict, Ridge, float, dict]:
    dataloader = DataLoader(
        dataset,
        batch_size=batch_size,
        shuffle=False,
        num_workers=num_workers,
        pin_memory=resolve_eval_pin_memory(args=args, default=False),
    )

    features, scores = compute_fr_features(
        model, dataloader, device, eval_type, dtype=torch.float64
    )

    return evaluate_ridge_metrics(
        dataset=dataset,
        features=features,
        scores=scores,
        num_splits=num_splits,
        phase=phase,
        alpha=alpha,
        grid_search=grid_search,
        args=args,
    )


def get_features_scores(
    model: nn.Module,
    dataloader: DataLoader,
    device: torch.device,
    eval_type: str = "scratch",
    dtype=torch.float32,
) -> Tuple[np.ndarray, np.ndarray]:
    """
    Get the features and scores for the given model and dataloader.

    Args:
        model (torch.nn.Module): model to test
        dataloader (torch.utils.data.Dataloader): dataloader
        device (torch.device): device to use for testing
        eval_type (str): Evaluation mode. This release supports only `scratch`.

    Returns:
        features (np.ndarray): features
        scores (np.ndarray): ground-truth MOS scores
    """
    num_samples = len(dataloader.dataset) * 5

    feat_dim = model.encoder.feat_dim
    proj_dim = model.encoder.projector[-1].out_features

    reps = torch.empty((num_samples, feat_dim * 2), dtype=dtype, device=device)
    embs = torch.empty((num_samples, proj_dim * 2), dtype=dtype, device=device)
    scores = torch.empty((num_samples,), dtype=dtype, device="cpu")

    idx = 0

    for i, batch in enumerate(dataloader):
        img_orig = batch["img"].to(device)
        img_ds = batch["img_ds"].to(device)
        mos = batch["mos"]
        img_orig = rearrange(img_orig, "b n c h w -> (b n) c h w")
        img_ds = rearrange(img_ds, "b n c h w -> (b n) c h w")
        mos = mos.repeat_interleave(5)

        amp_context = (
            torch.autocast(device_type="cuda")
            if device.type == "cuda" and torch.cuda.is_available()
            else nullcontext()
        )
        with amp_context, torch.no_grad():
            if eval_type == "scratch":
                f_orig, g_orig = model(img_orig)
                f_ds, g_ds = model(img_ds)
                f = torch.cat((f_orig, f_ds), dim=1)
                g = torch.cat((g_orig, g_ds), dim=1)
            else:
                raise ValueError(f"Eval type {eval_type} not supported")

        batch_size = f.size(0)

        reps[idx : idx + batch_size].copy_(f)
        embs[idx : idx + batch_size].copy_(g)
        scores[idx : idx + batch_size].copy_(mos)

        idx += batch_size

    reps = reps.cpu().numpy()
    embs = embs.cpu().numpy()
    scores = scores.numpy()

    return reps, embs, scores


def alpha_grid_search_1(
    dataset: Dataset, features: np.ndarray, scores: np.ndarray, num_splits: int
) -> float:

    grid_search_range = [1e-3, 1e3, 100]
    alphas = np.geomspace(*grid_search_range, endpoint=True)
    srocc_all = [[] for _ in range(len(alphas))]

    for i in range(num_splits):
        train_indices = dataset.get_split_indices(split=i, phase="train")
        val_indices = dataset.get_split_indices(split=i, phase="val")

        train_indices = np.repeat(train_indices * 5, 5) + np.tile(
            np.arange(5), len(train_indices)
        )
        val_indices = np.repeat(val_indices * 5, 5) + np.tile(
            np.arange(5), len(val_indices)
        )

        train_indices = train_indices[train_indices < features.shape[0]]
        val_indices = val_indices[val_indices < features.shape[0]]

        train_features = features[train_indices]
        train_scores = scores[train_indices]
        val_features = features[val_indices]
        val_scores = scores[val_indices]
        val_scores = val_scores[::5]

        if val_features.shape[0] < 1 or train_features.shape[0] < 1:
            continue

        for idx, alpha in enumerate(alphas):
            regressor = Ridge(alpha=alpha).fit(train_features, train_scores)

            preds = regressor.predict(val_features)
            preds = np.mean(
                np.reshape(preds, (-1, 5)), 1
            )  # Average the predictions of the 5 crops of the same image
            srocc_all[idx].append(stats.spearmanr(preds, val_scores)[0])

    srocc_all_median = [np.median(srocc) for srocc in srocc_all]
    srocc_all_median = np.array(srocc_all_median)
    best_alpha_idx = np.argmax(srocc_all_median)
    best_alpha_overall = alphas[best_alpha_idx]
    return (
        best_alpha_overall,
        best_alpha_idx,
        srocc_all_median,
        srocc_all_median[best_alpha_idx],
    )


def compute_srocc_for_split(
    train_fea, train_scr, val_fea, val_scr, coarse_alphas, fine_factor
):
    model = RidgeCV(alphas=coarse_alphas)
    model.fit(train_fea, train_scr)
    best_coarse_alpha = model.alpha_

    fine_alphas = np.geomspace(
        best_coarse_alpha / fine_factor, best_coarse_alpha * fine_factor, 10
    )

    XtX = train_fea.T @ train_fea
    Xty = train_fea.T @ train_scr
    eigvals, eigvecs = np.linalg.eigh(XtX)
    QtXty = eigvecs.T @ Xty

    sroccs = []
    for alpha in fine_alphas:
        inv = 1.0 / (eigvals + alpha)
        beta = eigvecs @ (inv * QtXty)
        preds = val_fea @ beta
        preds = preds.reshape(-1, 5).mean(axis=1)
        sroccs.append(stats.spearmanr(preds, val_scr)[0])

    best_idx = np.argmax(sroccs)
    return fine_alphas[best_idx], sroccs[best_idx]


def alpha_grid_search_11(dataset, features, scores, num_splits: int, n_jobs=8):
    coarse_alphas = np.geomspace(1e-3, 1e3, 10)
    fine_search_factor = 2

    tasks = []
    for i in range(num_splits):
        train_idx = dataset.get_split_indices(i, "train")
        val_idx = dataset.get_split_indices(i, "val")
        train_idx = np.repeat(train_idx * 5, 5) + np.tile(np.arange(5), len(train_idx))
        val_idx = np.repeat(val_idx * 5, 5) + np.tile(np.arange(5), len(val_idx))
        train_idx = train_idx[train_idx < features.shape[0]]
        val_idx = val_idx[val_idx < features.shape[0]]

        if len(train_idx) == 0 or len(val_idx) == 0:
            continue

        train_fea = features[train_idx]
        train_scr = scores[train_idx]
        val_fea = features[val_idx]
        val_scr = scores[val_idx][::5]

        tasks.append((train_fea, train_scr, val_fea, val_scr))

    results = Parallel(n_jobs=n_jobs)(
        delayed(compute_srocc_for_split)(*args, coarse_alphas, fine_search_factor)
        for args in tasks
    )

    if not results:
        return None

    best_alphas, sroccs = zip(*results)
    median_idx = np.argsort(sroccs)[len(sroccs) // 2]
    return best_alphas[median_idx], median_idx, list(sroccs), sroccs[median_idx]


def _safe_torch_load(checkpoint_path: Path) -> Mapping:
    """Load checkpoint with best-effort support for newer torch.load kwargs."""

    load_kwargs = {"map_location": "cpu"}
    try:
        if "weights_only" in inspect.signature(torch.load).parameters:
            load_kwargs["weights_only"] = True
    except (TypeError, ValueError):
        pass

    try:
        checkpoint = torch.load(checkpoint_path, **load_kwargs)
    except TypeError:
        load_kwargs.pop("weights_only", None)
        checkpoint = torch.load(checkpoint_path, **load_kwargs)

    return checkpoint


def _load_pretrained_weights(model: nn.Module, checkpoint_path: Path) -> None:
    """
    Load checkpoints saved either as a plain state_dict or wrapped inside a dict.
    Legacy pretraining runs stored the encoder weights under 'model.' / 'projector.' prefixes,
    so remap them to the 'encoder.' namespace expected by SimCLR.
    """

    checkpoint = _safe_torch_load(checkpoint_path)

    state_dict = checkpoint
    if isinstance(state_dict, Mapping):
        for key in ("state_dict", "model_state_dict", "model", "network"):
            candidate = state_dict.get(key) if isinstance(state_dict, Mapping) else None
            if isinstance(candidate, Mapping):
                state_dict = candidate
                break

    if not isinstance(state_dict, Mapping):
        raise RuntimeError(
            f"Unsupported checkpoint structure in {checkpoint_path}: expected a state_dict mapping."
        )

    normalized_state_dict = {}
    for key, value in state_dict.items():
        if key.startswith("module."):
            key = key[len("module.") :]
        normalized_state_dict[key] = value

    if not any(k.startswith("encoder.") for k in normalized_state_dict):
        remapped_state_dict = {}
        for key, value in normalized_state_dict.items():
            if key.startswith("model.") or key.startswith("projector."):
                remapped_state_dict[f"encoder.{key}"] = value
            else:
                remapped_state_dict[key] = value
        normalized_state_dict = remapped_state_dict

    model.load_state_dict(normalized_state_dict, strict=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=str, required=True, help="Configuration file")
    parser.add_argument(
        "--eval_type",
        type=str,
        default="scratch",
        choices=["scratch"],
        help="Evaluation type. Must be 'scratch' for this release.",
    )
    args, _ = parser.parse_known_args()
    eval_type = args.eval_type
    config = parse_config(args.config)
    args = parse_command_line_args(config)
    args = merge_configs(config, args)
    args.eval_type = eval_type
    args.data_base_path = Path(args.data_base_path)
    args = replace_none_string(args)
    if "logging" not in args:
        args.logging = DotMap(_dynamic=True)
    if "use_wandb" not in args.logging:
        args.logging.use_wandb = True
    if "wandb" not in args.logging:
        args.logging.wandb = DotMap(_dynamic=True)
    if "online" not in args.logging.wandb:
        args.logging.wandb.online = True
    if "test" not in args:
        args.test = DotMap(_dynamic=True)
    if "plcc_logistic" in args and "plcc_logistic" not in args.test:
        args.test.plcc_logistic = args.plcc_logistic

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    
    if args.eval_type != "scratch":
        raise ValueError(f"Eval type {args.eval_type} not supported")

    if args.model.method == "simclr":
        model = SimCLR(
            encoder_params=args.model.encoder, temperature=args.model.temperature
        )
    elif args.model.method == "vicreg":
        model = Vicreg(args)
    else:
        assert False

    checkpoint_base_path = PROJECT_ROOT / "experiments"
    assert (
        checkpoint_base_path / args.experiment_name
    ).exists(), (
        f"Experiment {(checkpoint_base_path / args.experiment_name)} does not exist"
    )
    checkpoint_path = checkpoint_base_path / args.experiment_name / "pretrain"
    checkpoint_path = [
        ckpt_path
        for ckpt_path in checkpoint_path.glob("*.pth")
        if "best" in ckpt_path.name
    ][0]
    _load_pretrained_weights(model, checkpoint_path)
    model.to(device)
    model.eval()

    if args.logging.use_wandb:
        wandb_config = prepare_wandb_config(args)
        logger = wandb.init(
            project=args.logging.wandb.project,
            entity=args.logging.wandb.entity,
            name=args.experiment_name,
            config=wandb_config,
            mode="online" if args.logging.wandb.online else "offline",
            resume="never",
        )
    else:
        logger = None

    test(args, model, logger, device)
